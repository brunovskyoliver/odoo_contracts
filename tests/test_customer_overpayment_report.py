# Copyright 2026 NOVEM IT
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

from datetime import datetime
import io
from unittest.mock import patch

from openpyxl import load_workbook

from odoo import Command, fields
from odoo.addons.account.tests.common import AccountTestInvoicingCommon


class TestCustomerOverpaymentReport(AccountTestInvoicingCommon):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.MoveLine = cls.env["account.move.line"]
        cls.MailMail = cls.env["mail.mail"]
        cls.company = cls.env.company
        cls.company.email = "company@example.com"
        cls.misc_journal = cls.company_data["default_journal_misc"]
        cls.bank_journal = cls.company_data["default_journal_bank"]
        cls.receivable_account = cls.company_data["default_account_receivable"]
        cls.revenue_account = cls.company_data["default_account_revenue"]
        cls.expense_account = cls.company_data["default_account_expense"]

    def _create_entry(self, partner, amount, date="2026-05-01", posted=True):
        if amount < 0:
            receivable_vals = {
                "name": "Customer overpayment",
                "partner_id": partner.id if partner else False,
                "account_id": self.receivable_account.id,
                "credit": abs(amount),
            }
            counterpart_vals = {
                "name": "Counterpart",
                "account_id": self.expense_account.id,
                "debit": abs(amount),
            }
        else:
            receivable_vals = {
                "name": "Customer invoice balance",
                "partner_id": partner.id if partner else False,
                "account_id": self.receivable_account.id,
                "debit": amount,
            }
            counterpart_vals = {
                "name": "Counterpart",
                "account_id": self.revenue_account.id,
                "credit": amount,
            }

        move = self.env["account.move"].create({
            "move_type": "entry",
            "journal_id": self.misc_journal.id,
            "date": date,
            "line_ids": [
                Command.create(receivable_vals),
                Command.create(counterpart_vals),
            ],
        })
        if posted:
            move.action_post()
        return move.line_ids.filtered(
            lambda line: line.account_id == self.receivable_account
        )

    def _create_non_receivable_negative_line(self, partner):
        move = self.env["account.move"].create({
            "move_type": "entry",
            "journal_id": self.misc_journal.id,
            "date": "2026-05-01",
            "line_ids": [
                Command.create({
                    "name": "Negative non-receivable",
                    "partner_id": partner.id,
                    "account_id": self.revenue_account.id,
                    "credit": 8,
                }),
                Command.create({
                    "name": "Counterpart",
                    "account_id": self.expense_account.id,
                    "debit": 8,
                }),
            ],
        })
        move.action_post()
        return move

    def _create_invoice(
        self,
        partner,
        payment_reference="FAK/2026/09001",
        amount=100,
        posted=True,
        invoice_date="2026-05-01",
    ):
        move = self.env["account.move"].create({
            "move_type": "out_invoice",
            "partner_id": partner.id,
            "invoice_date": invoice_date,
            "invoice_line_ids": [
                Command.create({
                    "name": "Invoice history",
                    "account_id": self.revenue_account.id,
                    "quantity": 1,
                    "price_unit": amount,
                }),
            ],
        })
        if posted:
            move.action_post()
            move.write({
                "payment_reference": payment_reference,
                "ref": payment_reference,
            })
        else:
            move.write({
                "payment_reference": payment_reference,
                "ref": payment_reference,
        })
        return move

    def _create_refund(
        self,
        invoice,
        amount=False,
        invoice_date="2026-05-02",
    ):
        refund = self.env["account.move"].create({
            "move_type": "out_refund",
            "partner_id": invoice.partner_id.id,
            "invoice_date": invoice_date,
            "reversed_entry_id": invoice.id,
            "ref": "Obrat: %s" % invoice.name,
            "invoice_line_ids": [
                Command.create({
                    "name": "Invoice refund",
                    "account_id": self.revenue_account.id,
                    "quantity": 1,
                    "price_unit": amount or invoice.amount_total,
                }),
            ],
        })
        refund.action_post()
        return refund

    def _create_partner_bank(self, partner, iban="SK3409000000005128707454"):
        return self.env["res.partner.bank"].create({
            "acc_number": iban,
            "partner_id": partner.id,
        })

    def _create_bank_statement_line(
        self,
        amount,
        payment_ref,
        partner=False,
        transaction_details=False,
        date="2026-05-10",
    ):
        values = {
            "amount": amount,
            "date": date,
            "payment_ref": payment_ref,
            "journal_id": self.bank_journal.id,
        }
        if partner:
            values["partner_id"] = partner.id
        if transaction_details:
            values["transaction_details"] = transaction_details
        return self.env["account.bank.statement.line"].create(values)

    def _reconcile_statement_line_with_invoice(self, statement_line, invoice):
        _liquidity_lines, suspense_lines, _other_lines = (
            statement_line.with_context(
                skip_account_move_synchronization=True
            )._seek_for_lines()
        )
        invoice_line = invoice.line_ids.filtered(
            lambda line: line.account_id.account_type == "asset_receivable"
        )
        suspense_lines.with_context(check_move_validity=False).write({
            "account_id": invoice_line.account_id.id,
            "partner_id": invoice.partner_id.id,
        })
        (suspense_lines + invoice_line).reconcile()
        return statement_line

    def test_due_check_uses_braslava_time(self):
        self.assertTrue(
            self.MoveLine._is_customer_overpayment_report_due(
                now_utc=datetime(2026, 6, 15, 8, 0, 0)
            )
        )
        self.assertFalse(
            self.MoveLine._is_customer_overpayment_report_due(
                now_utc=datetime(2026, 6, 14, 8, 0, 0)
            )
        )
        self.assertFalse(
            self.MoveLine._is_customer_overpayment_report_due(
                now_utc=datetime(2026, 6, 15, 7, 0, 0)
            )
        )

    def test_outside_due_window_skips_without_search_or_mail(self):
        with patch.object(
            type(self.MoveLine), "_get_customer_overpayment_report_data"
        ) as report_data, patch.object(
            type(self.MailMail), "send", return_value=True
        ) as mail_send:
            result = self.MoveLine._check_and_send_customer_overpayment_report(
                now_utc=datetime(2026, 6, 15, 7, 0, 0)
            )

        self.assertFalse(result)
        self.assertFalse(report_data.called)
        self.assertFalse(mail_send.called)

    def test_manual_cron_trigger_bypasses_due_window(self):
        with patch.object(
            type(self.MoveLine),
            "_get_customer_overpayment_report_data",
            return_value=[{
                "partner": self.partner_a,
                "lines": self.env["account.move.line"],
                "bank_payments": [],
                "overpaid_amount": 12,
                "receivable_overpaid_amount": 12,
                "bank_overpaid_amount": 0,
                "total_balance": -12,
                "line_count": 0,
            }],
        ), patch.object(type(self.MailMail), "send", return_value=True) as mail_send:
            result = self.MoveLine.with_context(
                lastcall=datetime(2026, 6, 1, 8, 0, 0)
            )._check_and_send_customer_overpayment_report(
                now_utc=datetime(2026, 6, 3, 8, 0, 0)
            )

        self.assertTrue(result)
        self.assertEqual(mail_send.call_count, 1)

    def test_report_includes_any_credit_customer_with_full_open_context(self):
        overpaid_partner = self.env["res.partner"].create({
            "name": "Any Credit Customer",
        })
        positive_partner = self.env["res.partner"].create({
            "name": "Positive Only Customer",
        })
        excluded_partner = self.env["res.partner"].create({
            "name": "Excluded Customer",
        })
        no_partner_line = self._create_entry(None, -2)
        credit_line = self._create_entry(overpaid_partner, -15)
        positive_context_line = self._create_entry(overpaid_partner, 40)
        self._create_entry(positive_partner, 30)
        self._create_entry(excluded_partner, -20, posted=False)
        reconciled_credit = self._create_entry(excluded_partner, -5)
        reconciled_debit = self._create_entry(excluded_partner, 5)
        (reconciled_credit + reconciled_debit).reconcile()
        self._create_non_receivable_negative_line(excluded_partner)

        report_data = self.MoveLine._get_customer_overpayment_report_data()

        self.assertEqual(len(report_data), 1)
        self.assertEqual(report_data[0]["partner"], overpaid_partner)
        self.assertEqual(report_data[0]["lines"], credit_line + positive_context_line)
        self.assertEqual(report_data[0]["overpaid_amount"], 15)
        self.assertEqual(report_data[0]["receivable_overpaid_amount"], 15)
        self.assertEqual(report_data[0]["bank_overpaid_amount"], 0)
        self.assertEqual(report_data[0]["total_balance"], 25)
        self.assertNotIn(no_partner_line, report_data[0]["lines"])

        body = self.MoveLine._render_customer_overpayment_report_email(
            report_data,
            fields.Date.from_string("2026-06-15"),
        )
        self.assertIn("Any Credit Customer", body)
        self.assertNotIn("Počet otvorených riadkov", body)
        self.assertNotIn("Customer overpayment", body)
        self.assertNotIn("Customer invoice balance", body)
        self.assertNotIn("Positive Only Customer", body)
        self.assertNotIn("Excluded Customer", body)
        workbook = load_workbook(
            io.BytesIO(
                self.MoveLine._generate_customer_overpayment_report_xlsx(
                    report_data,
                    fields.Date.from_string("2026-06-15"),
                )
            )
        )
        self.assertIn("Any Credit Customer", workbook.sheetnames)
        sheet = workbook["Any Credit Customer"]
        values = [
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if cell.value
        ]
        self.assertIn("Customer overpayment", values)
        self.assertIn("Customer invoice balance", values)

    def test_report_includes_bank_only_customer_by_registered_iban(self):
        partner = self.env["res.partner"].create({"name": "Bank Only Customer"})
        self._create_invoice(partner, payment_reference="FAK/2026/09002")
        self._create_partner_bank(partner, "SK8583300000002001378223")
        statement_line = self._create_bank_statement_line(
            42,
            "Unexpected payment, SK8583300000002001378223",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()

        partner_report = [
            data for data in report_data if data["partner"] == partner
        ]
        self.assertEqual(len(partner_report), 1)
        self.assertEqual(partner_report[0]["overpaid_amount"], 42)
        self.assertEqual(partner_report[0]["bank_overpaid_amount"], 42)
        self.assertEqual(
            partner_report[0]["bank_payments"][0]["statement_line"],
            statement_line,
        )

    def test_bank_payment_merges_with_receivable_credit(self):
        partner = self.env["res.partner"].create({"name": "Merged Customer"})
        self._create_invoice(partner, payment_reference="FAK/2026/09003")
        self._create_partner_bank(partner, "SK8702000000003449448458")
        self._create_entry(partner, -15)
        self._create_bank_statement_line(
            20,
            "Extra payment SK8702000000003449448458",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertEqual(partner_report["overpaid_amount"], 35)
        self.assertEqual(partner_report["receivable_overpaid_amount"], 15)
        self.assertEqual(partner_report["bank_overpaid_amount"], 20)

    def test_bank_payment_xlsx_includes_refund_transfer_helpers(self):
        partner = self.env["res.partner"].create({"name": "Refund Helper Customer"})
        invoice = self._create_invoice(
            partner,
            payment_reference="FAK/2026/09012",
            amount=120.48,
            invoice_date="2026-05-23",
        )
        invoice.write({"name": "FAK/2026/09012"})
        self._create_refund(invoice, amount=120.48, invoice_date="2026-05-25")
        self._create_partner_bank(partner, "SK5256000000007889942001")
        self._create_bank_statement_line(
            120.48,
            "REFUND HELPER CUSTOMER, SK5256000000007889942001",
            date="2026-05-27",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        workbook = load_workbook(
            io.BytesIO(
                self.MoveLine._generate_customer_overpayment_report_xlsx(
                    report_data,
                    fields.Date.from_string("2026-06-15"),
                )
            )
        )
        sheet = workbook["Refund Helper Customer"]
        headers = [sheet.cell(8, column).value for column in range(1, 13)]
        values = [
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if cell.value
        ]

        self.assertEqual(headers[-2:], ["Poznámka", "VS"])
        self.assertIn("Vrátenie mylnej platby k FAK/2026/09012", values)
        self.assertIn("202609012", values)

    def test_refund_transfer_helpers_ignore_invoices_older_than_month(self):
        partner = self.env["res.partner"].create({"name": "Old Refund Customer"})
        invoice = self._create_invoice(
            partner,
            payment_reference="FAK/2026/09013",
            amount=120.48,
            invoice_date="2026-04-20",
        )
        invoice.write({"name": "FAK/2026/09013"})
        self._create_refund(invoice, amount=120.48, invoice_date="2026-04-22")
        self._create_partner_bank(partner, "SK0309000000005223957259")
        self._create_bank_statement_line(
            120.48,
            "OLD REFUND CUSTOMER, SK0309000000005223957259",
            date="2026-05-27",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertFalse(partner_report["bank_payments"][0]["refund_note"]["text"])
        self.assertFalse(
            partner_report["bank_payments"][0]["refund_note"]["variable_symbol"]
        )

    def test_refund_transfer_helpers_allow_refund_after_payment_date(self):
        partner = self.env["res.partner"].create({"name": "Late Refund Customer"})
        invoice = self._create_invoice(
            partner,
            payment_reference="FAK/2026/09014",
            amount=76.64,
            invoice_date="2026-05-23",
        )
        invoice.write({"name": "FAK/2026/09014"})
        self._create_refund(invoice, amount=76.64, invoice_date="2026-05-25")
        self._create_partner_bank(partner, "SK0575000000004035223042")
        self._create_bank_statement_line(
            76.64,
            "LATE REFUND CUSTOMER, SK0575000000004035223042",
            date="2026-05-24",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertEqual(
            partner_report["bank_payments"][0]["refund_note"]["text"],
            "Vrátenie mylnej platby k FAK/2026/09014",
        )
        self.assertEqual(
            partner_report["bank_payments"][0]["refund_note"]["variable_symbol"],
            "202609014",
        )

    def test_refund_transfer_helpers_prefer_paid_invoice_reference(self):
        partner = self.env["res.partner"].create({"name": "Named Invoice Customer"})
        older_invoice = self._create_invoice(
            partner,
            payment_reference="FAK/2026/09015",
            amount=76.64,
            invoice_date="2026-05-22",
        )
        older_invoice.write({"name": "FAK/2026/09015"})
        named_invoice = self._create_invoice(
            partner,
            payment_reference="FAK/2026/09016",
            amount=76.64,
            invoice_date="2026-05-23",
        )
        named_invoice.write({"name": "FAK/2026/09016"})
        self._create_refund(older_invoice, amount=76.64, invoice_date="2026-05-24")
        self._create_refund(named_invoice, amount=76.64, invoice_date="2026-05-25")
        self._create_bank_statement_line(
            76.64,
            "Uhrada faktury FAK/2026/09015",
            date="2026-05-24",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertEqual(
            partner_report["bank_payments"][0]["refund_note"]["text"],
            "Vrátenie mylnej platby k FAK/2026/09015",
        )
        self.assertEqual(
            partner_report["bank_payments"][0]["refund_note"]["variable_symbol"],
            "202609015",
        )

    def test_positive_only_customer_excluded_without_bank_payment(self):
        positive_partner = self.env["res.partner"].create({
            "name": "Positive Without Bank Customer",
        })
        self._create_invoice(positive_partner, payment_reference="FAK/2026/09004")

        report_data = self.MoveLine._get_customer_overpayment_report_data()

        self.assertNotIn(
            positive_partner,
            [data["partner"] for data in report_data],
        )

    def test_bank_payment_matching_by_invoice_reference(self):
        partner = self.env["res.partner"].create({"name": "Invoice Ref Customer"})
        self._create_invoice(partner, payment_reference="FAK/2026/09005")
        self._create_bank_statement_line(
            33,
            "Uhrada faktury FAK/2026/09005",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertEqual(partner_report["bank_overpaid_amount"], 33)
        self.assertEqual(
            partner_report["bank_payments"][0]["match_method"],
            "Referencia faktúry",
        )

    def test_bank_payment_matching_by_historical_iban(self):
        partner = self.env["res.partner"].create({"name": "Historical Customer"})
        invoice = self._create_invoice(
            partner,
            payment_reference="FAK/2026/09006",
            amount=25,
        )
        historical_line = self._create_bank_statement_line(
            25,
            "Historical payment SK0575000000004035223042",
            partner=partner,
        )
        self._reconcile_statement_line_with_invoice(historical_line, invoice)
        self.assertTrue(historical_line.is_reconciled)
        self._create_bank_statement_line(
            7,
            "New unmatched payment SK0575000000004035223042",
        )

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertEqual(partner_report["bank_overpaid_amount"], 7)
        self.assertEqual(
            partner_report["bank_payments"][0]["match_method"],
            "Historické párovanie IBAN",
        )

    def test_bank_payment_exclusions(self):
        matched_partner = self.env["res.partner"].create({
            "name": "Matched Exclusion Customer",
        })
        no_history_partner = self.env["res.partner"].create({
            "name": "No History Customer",
        })
        draft_partner = self.env["res.partner"].create({
            "name": "Draft Invoice Customer",
        })
        reconciled_partner = self.env["res.partner"].create({
            "name": "Reconciled Customer",
        })
        self._create_invoice(matched_partner, payment_reference="FAK/2026/09007")
        self._create_partner_bank(no_history_partner, "SK7565000000000016603949")
        self._create_invoice(
            draft_partner,
            payment_reference="FAK/2026/09008",
            posted=False,
        )
        reconciled_invoice = self._create_invoice(
            reconciled_partner,
            payment_reference="FAK/2026/09009",
            amount=19,
        )
        reconciled_line = self._create_bank_statement_line(
            19,
            "FAK/2026/09009",
            partner=reconciled_partner,
        )
        self._reconcile_statement_line_with_invoice(
            reconciled_line,
            reconciled_invoice,
        )
        self._create_bank_statement_line(
            9,
            "No history SK7565000000000016603949",
        )
        self._create_bank_statement_line(11, "FAK/2026/09008")
        self._create_bank_statement_line(-13, "FAK/2026/09007")
        self._create_bank_statement_line(14, "No match")

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partners = [data["partner"] for data in report_data]

        self.assertNotIn(no_history_partner, partners)
        self.assertNotIn(draft_partner, partners)
        self.assertNotIn(reconciled_partner, partners)
        self.assertNotIn(matched_partner, partners)

    def test_ambiguous_bank_payment_match_is_excluded(self):
        partner_a = self.env["res.partner"].create({"name": "Ambiguous A"})
        partner_b = self.env["res.partner"].create({"name": "Ambiguous B"})
        self._create_invoice(partner_a, payment_reference="FAK/2026/09010")
        self._create_invoice(partner_b, payment_reference="FAK/2026/09010")
        self._create_bank_statement_line(51, "FAK/2026/09010")

        report_data = self.MoveLine._get_customer_overpayment_report_data()

        self.assertNotIn(partner_a, [data["partner"] for data in report_data])
        self.assertNotIn(partner_b, [data["partner"] for data in report_data])

    def test_bank_payment_backed_by_receivable_credit_is_not_double_counted(self):
        partner = self.env["res.partner"].create({"name": "Deduped Customer"})
        self._create_invoice(partner, payment_reference="FAK/2026/09011")
        statement_line = self._create_bank_statement_line(
            25,
            "FAK/2026/09011",
            partner=partner,
        )
        _liquidity_lines, suspense_lines, _other_lines = (
            statement_line.with_context(
                skip_account_move_synchronization=True
            )._seek_for_lines()
        )
        suspense_lines.with_context(check_move_validity=False).write({
            "account_id": self.receivable_account.id,
            "partner_id": partner.id,
        })

        report_data = self.MoveLine._get_customer_overpayment_report_data()
        partner_report = [
            data for data in report_data if data["partner"] == partner
        ][0]

        self.assertEqual(partner_report["overpaid_amount"], 25)
        self.assertEqual(partner_report["receivable_overpaid_amount"], 25)
        self.assertEqual(partner_report["bank_overpaid_amount"], 0)

    def test_currency_zero_overpayments_are_excluded(self):
        partner = self.env["res.partner"].create({"name": "Tiny Credit Customer"})
        self._create_entry(partner, -0.001)

        report_data = self.MoveLine._get_customer_overpayment_report_data()

        self.assertFalse(report_data)

    def test_due_window_without_results_sends_no_email(self):
        with patch.object(
            type(self.MoveLine),
            "_get_customer_overpayment_report_data",
            return_value=[],
        ), patch.object(type(self.MailMail), "send", return_value=True) as mail_send:
            result = self.MoveLine._check_and_send_customer_overpayment_report(
                now_utc=datetime(2026, 6, 15, 8, 0, 0)
            )

        self.assertFalse(result)
        self.assertFalse(mail_send.called)

    def test_due_window_with_results_sends_report_email(self):
        partner = self.env["res.partner"].create({"name": "Mail Customer"})
        self._create_entry(partner, -12)
        self.env["ir.config_parameter"].sudo().set_param(
            "web.base.url",
            "https://odoo.example.test",
        )

        with patch.object(type(self.MailMail), "send", return_value=True) as mail_send:
            result = self.MoveLine._check_and_send_customer_overpayment_report(
                now_utc=datetime(2026, 6, 15, 8, 0, 0)
            )

        self.assertTrue(result)
        self.assertEqual(mail_send.call_count, 1)
        mail = self.MailMail.search(
            [("subject", "=", "Prehľad preplatkov zákazníkov - 2026-06-15")],
            order="id desc",
            limit=1,
        )
        self.assertTrue(mail)
        self.assertEqual(
            mail.email_to,
            self.MoveLine._CUSTOMER_OVERPAYMENT_REPORT_RECIPIENTS,
        )
        self.assertIn("Mail Customer", mail.body_html)
        self.assertIn(
            (
                'href="https://odoo.example.test/web#id=%s&amp;'
                'model=res.partner&amp;view_type=form"'
            ) % partner.id,
            mail.body_html,
        )
        self.assertIn('target="_blank"', mail.body_html)
        self.assertNotIn("Customer overpayment", mail.body_html)
        self.assertEqual(len(mail.attachment_ids), 1)
        self.assertEqual(
            mail.attachment_ids.mapped("mimetype"),
            [
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ],
        )

    def test_cron_xml_is_configured(self):
        cron = self.env.ref("contract.ir_cron_customer_overpayment_report")

        self.assertEqual(cron.model_id.model, "account.move.line")
        self.assertEqual(cron.code, "model.cron_send_customer_overpayment_report()")
        self.assertEqual(cron.interval_number, 1)
        self.assertEqual(cron.interval_type, "days")
        self.assertEqual(
            fields.Datetime.to_string(cron.nextcall),
            "2026-06-15 08:00:00",
        )
