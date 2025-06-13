import sys
import tkinter as tk
from tkinter import ttk, scrolledtext
import pandas as pd
from tkinter import messagebox
import os
from typing import Callable
import threading
import tkinter.filedialog as filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import re
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
def get_unique_o2_services(file: str):
    try:
        df = pd.read_csv(file, encoding="utf-8", low_memory=False)
        # Get all recurring fees except VPN services
        recurring_services = df[
            (df['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeType'] == 'recurring_arrears') &
            (~df['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName'].str.contains('VPN', na=False))
        ]['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName'].unique()
        
        print("\nUnique O2 recurring services across all numbers:")
        print("-" * 50)
        for service in sorted(recurring_services):
            if pd.notna(service):  # Only print non-null services
                print(f"{service}")
            
        return recurring_services
        
    except Exception as e:
        print(f"Error processing O2 CSV: {str(e)}")
        return []
def get_unique_tbiznis_services(file: str):
    try:
        df = pd.read_csv(file, encoding="utf-8", low_memory=False)
        # Get all rows containing T-Biznis services
        tbiznis_services = df[
            df['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('T-Biznis', na=False)
        ]['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].unique()
        
        print("\nUnique T-Biznis services across all numbers:")
        print("-" * 50)
        for service in sorted(tbiznis_services):
            print(f"{service}")
            
        return tbiznis_services
        
    except Exception as e:
        print(f"Error processing Telekom CSV: {str(e)}")
        return []


def preprocess_o2_csv(file_path: str) -> str:
    """
    Preprocess O2 CSV file to fill in missing phone numbers.
    Returns path to the preprocessed file.
    """
    # Read the raw CSV
    df = pd.read_csv(file_path, encoding="utf-8", low_memory=False)
    
    # Get the column index for MSISDN (phone number)
    msisdn_col = [col for col in df.columns if 'MSISDN' in col][0]
    
    # Forward fill the phone numbers
    current_number = None
    processed_rows = []
    
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        
        # Update current number if we find a non-empty value
        if pd.notna(row[msisdn_col]):
            current_number = row[msisdn_col]
        
        # Fill in the number if it's empty
        if current_number is not None:
            row_dict[msisdn_col] = current_number
            
        processed_rows.append(row_dict)
    
    # Create new DataFrame with filled numbers
    processed_df = pd.DataFrame(processed_rows)
    
    # Save to temporary file
    temp_file = file_path.replace('.csv', '_processed.csv')
    processed_df.to_csv(temp_file, index=False)
    
    return temp_file

def insert_image(sheet, image_path, cell):
    """Insert an image into the specified cell of the sheet."""
    try:
         # Load the image
        img = Image(image_path)
        
        # Adjust the size of the image if needed (optional)
        img.width, img.height = 200, 100  # Resize if necessary
        
        # Extract column letter and row number from cell reference
        col_letter = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell)))
        
        # Get column width in pixels (1 unit = 7 pixels)
        col_width = sheet.column_dimensions[col_letter].width
        col_width_px = col_width * 7
        
        # Calculate offsets to center the image
        x_offset = (col_width_px - img.width) / 2
        
        # Adjust row height to fit the image
        row_height = img.height * 0.75  # Convert to points (Excel units)
        sheet.row_dimensions[row_num].height = row_height
        
        # Calculate y_offset to center vertically
        y_offset = (row_height - img.height * 0.75) / 2
        
        # Add the image with calculated offsets
        img.anchor = cell
        img.left = x_offset
        img.top = y_offset
        
        sheet.add_image(img)
        
    except Exception as e:
        print(f"Error inserting image: {e}")


def adjust_column_width(sheet, max_width=100):
    """Adjust the width of columns in the sheet to fit the content, with a max limit."""
    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)  # Get column letter
        for cell in column_cells:
            try:
                if cell.value:  # Check if the cell has a value
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Adjust the column width with a max width limit
        adjusted_width = min(max_length, max_width)  # Add padding, but cap at max_width
        sheet.column_dimensions[column].width = adjusted_width

def clean_phone_number(phone_number: str) -> str:
    """Remove all non-numeric characters from a phone number and strip leading '00'."""
    # Remove all non-numeric characters
    cleaned_number = re.sub(r'\D', '', phone_number)
    # Strip leading '00'
    return cleaned_number.lstrip('00')

def print_vyrobne_cislo_column(customer_data):
    print(customer_data['number'])

def ensure_directory_exists(directory):
    """Create directory if it doesn't exist."""
    if not os.path.exists(directory):
        os.makedirs(directory)

def get_output_folder(customer_info: tuple, base_dir: str = "output") -> str:
    """Determine the appropriate output folder based on customer info."""
    customer_type, name = customer_info
    
    # Create base output directory if it doesn't exist
    ensure_directory_exists(base_dir)
    
    # Create companies and individuals directories
    companies_dir = os.path.join(base_dir, "companies")
    individuals_dir = os.path.join(base_dir, "individuals")
    ensure_directory_exists(companies_dir)
    ensure_directory_exists(individuals_dir)
    
    if customer_type == 'company':
        company_dir = os.path.join(companies_dir, name)
        ensure_directory_exists(company_dir)
        return company_dir
    else:  # 'individual' or 'unknown'
        ensure_directory_exists(individuals_dir)
        return individuals_dir

def load_customer_data(customer_file: str):
    """Load the customer data CSV into a DataFrame."""
    try:
        return pd.read_csv(customer_file, encoding="utf-8", low_memory=False)
    except Exception as e:
        print(f"Error loading customer data: {e}")
        return None

def get_customer_name_or_company(customer_data, phone_number):
    """Return the customer name or company based on the phone number."""
    cleaned_number = clean_phone_number(phone_number)
    
    customer_data['Výrobné číslo'] = customer_data['Výrobné číslo'].astype(str).str.strip()
    customer_row = customer_data[customer_data['Výrobné číslo'] == cleaned_number]
    
    if not customer_row.empty:
        company = customer_row['PO názov firmy'].iloc[0]
        first_name = customer_row['FO Meno'].iloc[0]
        last_name = customer_row['FO priezvisko'].iloc[0]
        
        # Return tuple with type identifier and name
        if pd.notna(company) and company != "":
            return ('company', company)
        
        if pd.notna(first_name) and pd.notna(last_name) and first_name != "" and last_name != "":
            return ('individual', f"{first_name} {last_name}")
    
    return ('unknown', "Unknown")


def format_duration(seconds):
    """Format seconds into HH:MM:SS format."""
    try:
        seconds = float(seconds)
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    except ValueError:
        return "00:00:00"

def format_data_usage(bytes_or_mb):
    """Format bytes or MB into appropriate unit (MB or GB)."""
    try:
        if bytes_or_mb > 1024 * 1024:
            mb = float(bytes_or_mb) / (1024 * 1024)
        else:
            mb = float(bytes_or_mb)
            
        # if mb >= 1024:
        #     return f"{mb/1024:.2f} GB"
        return f"{mb:.2f} MB"
    except ValueError:
        return "0"

def excel_to_pdf(excel_path: str) -> None:
    try:
        print(f"Starting PDF conversion for: {excel_path}")
        wb = load_workbook(excel_path)
        sheet = wb.active
        
        pdf_path = excel_path.replace('.xlsx', '.pdf')
        print(f"Will create PDF at: {pdf_path}")

        pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf'))
       
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        
        styles = getSampleStyleSheet()
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=8,
            leading=10,
            spaceBefore=0,
            spaceAfter=0,
            encoding='utf-8'
        )

        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Normal'],
            fontName='DejaVuSans-Bold',
            fontSize=10,
            leading=12,
            spaceBefore=0,
            spaceAfter=0,
            encoding='utf-8'
        )
        
        elements = []
        table_data = []
        header_processed = False
        column_header_style = ParagraphStyle(
            'ColumnHeader',
            parent=styles['Normal'],
            fontName='DejaVuSans-Bold',
            fontSize=8, 
            leading=10,
            spaceBefore=0,
            spaceAfter=0,
            encoding='utf-8'
        )
        section_headers = {
            "Rozpis účtovaných poplatkov": ["Popis", "DPH", "Suma v EUR bez DPH"],
            "Rozpis SMS / MMS": ["Popis", "Počet kusov", ""],
            "Rozpis volaní": ["Popis", "Trvanie hovorov", ""],
            "Rozpis dát": ["Popis", "Spotreba dát v MB", ""]
        }
        
        for row in sheet.rows:
            row_text = [str(cell.value) if cell.value is not None else "" for cell in row]
            
            if not any(text.strip() for text in row_text):
                continue
            
            if not header_processed:
                if "Rozpis spotreby pre telefónne číslo:" in row_text[0]:
                    phone_text = f"{row_text[0].strip()} {row_text[1].strip()}"
                    elements.append(Paragraph(phone_text, header_style))
                    elements.append(Spacer(1, 10))
                elif "Za účtovné obdobie" in row_text[0]:
                    period_text = f"{row_text[0].strip()} {row_text[1].strip()}"
                    elements.append(Paragraph(period_text, normal_style))
                    elements.append(Spacer(1, 10))
                    header_processed = True
                continue
            
            
            is_header = any(header in row_text[0] for header in section_headers.keys())
            
            if is_header:
            
                if table_data:
                    table = Table(table_data, colWidths=[300, 100, 100])
                    table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ]))
                    elements.append(table)
                    elements.append(Spacer(1, 20))  
                
                
                elements.append(Paragraph(row_text[0], header_style))
                elements.append(Spacer(1, 15))  
                
                
                table_data = []
                for section, headers in section_headers.items():
                    if section in row_text[0]:
                        header_row = [Paragraph(h, column_header_style) for h in headers]
                        table_data.append(header_row)
                        break
                
                continue
            
            
            row_data = []
            
            if "Faktúrovaná suma nad paušál bez DPH:" in row_text[0]:
                continue
                
            for i, cell_text in enumerate(row_text):
                if cell_text.strip():
                    p = Paragraph(cell_text, normal_style)
                    row_data.append(p)
                else:
                    row_data.append("")
            
            if row_data:
                table_data.append(row_data)
        
    
        if "Faktúrovaná suma nad paušál bez DPH:" in row_text[0]:
            
            if table_data:
                table = Table(table_data, colWidths=[300, 100, 100])
                table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 20))
            
        
            if "Faktúrovaná suma nad paušál bez DPH:" in row_text[0]:
                elements.append(Spacer(1, 20))
                fakturovana_data = [[
                    Paragraph("Faktúrovaná suma nad paušál bez DPH:", header_style),
                    Paragraph(row_text[1], normal_style) if row_text[1] else ""
                ]]
                fakturovana_table = Table(fakturovana_data, colWidths=[300, 100])
                fakturovana_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(fakturovana_table)
                table_data = []
        
        if table_data:
            table = Table(table_data, colWidths=[300, 100, 100])
            table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(table)
        
        # Build PDF
        print("Building PDF...")
        doc.build(elements)
        print(f"PDF created successfully at: {pdf_path}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error converting {excel_path} to PDF: {e}")

class OutputManager:
    
    def __init__(self, gui_callback: Callable[[str], None]):
        self.gui_callback = gui_callback
        self.current_txt_file = None
        self.current_xlsx_workbook = None
        self.current_number = None
        self.current_customer = None
        self.xlsx_filepath = None
        self.current_row = 1
        self.company_totals = {}
        self.current_company = None
    
    def start_new_file(self, number: str, customer_type_and_info: tuple):
        """Start writing to new txt and xlsx files for the given number."""
        #print(f"Starting new file for number: {number}")  # Debug print
        customer_type, customer_name = customer_type_and_info
        self.current_number = number
        self.current_customer = customer_name

        if customer_type == 'company':
            self.current_company = customer_name
            # Initialize company total if not exists
            if customer_name not in self.company_totals:
                self.company_totals[customer_name] = 0.0
        # Setup txt file
        txt_output_folder = get_output_folder(customer_type_and_info)
        timestamp = datetime.now().strftime("%Y%m%d")
        
        if customer_type == 'company':
            #txt_filename = f"{number}_{timestamp}.txt"
            txt_filename = f"{number}.txt"
            txt_filepath = os.path.join(txt_output_folder, txt_filename)
        else:
            # Handle individual customers
            try:
                # Safely attempt to split the customer name
                last_name = customer_name.split(' ')[1]
            except IndexError:
                # Default to 'unknown' if the name doesn't split properly
                last_name = 'unknown'
            
            txt_filename = f"{number}_{last_name}.txt"
            txt_filepath = os.path.join(txt_output_folder, txt_filename)
        # Setup xlsx file
        xlsx_output_folder = get_output_folder(customer_type_and_info, base_dir="output_excel")
        #print(f"Excel output folder: {xlsx_output_folder}")  # Debug print
        if customer_type == 'company':
            #xlsx_filename = f"{number}_{timestamp}.xlsx"
            xlsx_filename = f"{number}.xlsx"
            self.xlsx_filepath = os.path.join(xlsx_output_folder, xlsx_filename)
        else:
            # Handle individual customers
            try:
                # Safely attempt to split the customer name
                last_name = customer_name.split(' ')[1]
            except IndexError:
                # Default to 'unknown' if the name doesn't split properly
                last_name = 'unknown'
            xlsx_filename = f"{number}_{last_name}.xlsx"
            self.xlsx_filepath = os.path.join(xlsx_output_folder, xlsx_filename)
       # print(f"Excel filepath: {self.xlsx_filepath}")  # Debug print
        
        # Open txt file
        if self.current_txt_file is not None:
            self.current_txt_file.close()
        self.current_txt_file = open(txt_filepath, "w", encoding="utf-8")
        
        # Create new Excel workbook
        if self.current_xlsx_workbook is not None:
            try:
                self.current_xlsx_workbook.save(self.xlsx_filepath)
            except Exception as e:
                print(f"Error saving previous workbook: {e}")
            self.current_xlsx_workbook.close()
        
        self.current_xlsx_workbook = Workbook()
        self.current_sheet = self.current_xlsx_workbook.active
        self.current_row = 1
        #print("New workbook created")  # Debug print
        
        # Try to save empty workbook to verify path is writable
        try:
            self.current_xlsx_workbook.save(self.xlsx_filepath)
            #print("Initial Excel file saved successfully")  # Debug print
            image_path = "novem.png"  # Replace with your actual image path
            insert_image(self.current_sheet, image_path, "A1")
        except Exception as e:
            print(f"Error saving initial Excel file: {e}")

    
    def write(self, text: str):
        thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

        """Write text to GUI, txt file, and format for Excel."""
        # toto je celkom zle vyriesene priznavam sa, ale it is what it is
        if "T-Biznis Flex Variant 1" in text:
            text = text.replace("T-Biznis Flex Variant 1", "NOVEM nekonečno 6GB")
        if "T-Biznis Flex Variant 10" in text:
            text = text.replace("T-Biznis Flex Variant 10", "NOVEM nekonečno 10GB")
        if "T-Biznis Flex Variant 11" in text:
            text = text.replace("T-Biznis Flex Variant 11", "NOVEM nekonečno 30GB")
        if "T-Biznis Flex Variant 2" in text:
            text = text.replace("T-Biznis Flex Variant 2", "NOVEM Fér bez dát")
        if "T-Biznis Flex Variant 3" in text:
            text = text.replace("T-Biznis Flex Variant 3", "NOVEM nekonečno 20GB")
        if "T-Biznis Flex Variant 4" in text:
            text = text.replace("T-Biznis Flex Variant 4", "NOVEM nekonečno 50GB")
        if "T-Biznis Flex Variant 5" in text:
            text = text.replace("T-Biznis Flex Variant 5", "NOVEM 250 0,5GB")
        if "T-Biznis Flex Variant 6" in text:
            text = text.replace("T-Biznis Flex Variant 6", "NOVEM nekonečno 0,5GB")
        if "T-Biznis Flex Variant 7" in text:
            text = text.replace("T-Biznis Flex Variant 7", "NOVEM 250 30 GB")
        if "T-Biznis Flex Variant 8" in text:
            text = text.replace("T-Biznis Flex Variant 8", "NOVEM 250 10 GB")
        if "T-Biznis Flex Variant 9" in text:
            text = text.replace("T-Biznis Flex Variant 9", "NOVEM nekonečno bez dát")
        if "e-Net" in text:
            text = text.replace("e-Net", "NOVEM")
            if "minút" in text:
                text = text.replace("minút ", "")
            if "minut" in text:
                text = text.replace("minut ", "")
        if "TELEKOM TELEFONNE CISLO" in text:
            text = text.replace("TELEKOM TELEFONNE CISLO", "Rozpis spotreby pre telefónne číslo:")
        if "O2 TELEFONNE CISLO" in text:
            text = text.replace("O2 TELEFONNE CISLO", "Rozpis spotreby pre telefónne číslo:")
        if "(MB)" in text:
            text = text.replace(" (MB)", "")
        if "(GB)" in text:
            text = text.replace(" (GB)", "")
        match = False
        
        self.gui_callback(text)
        if self.current_txt_file is not None:
            self.current_txt_file.write(text + "\n")
            self.current_txt_file.flush()
        
        # tuto formatujeme excel
        if self.current_xlsx_workbook is not None:
            try:
                #print(text)
                if text.startswith("="):  # Hladame zaciatok a koniec tento if velmi moc nefunguje
                    if "TELEFONNE CISLO" in text: 
                        # Write header
                        self.current_sheet.merge_cells(f'A{self.current_row}:B{self.current_row}')
                        clean_text = text.strip("=").strip()
                        self.current_sheet.cell(self.current_row, 1, clean_text).font = Font(bold=True)
                    elif "Celková:" in text: # ani tento moc nefunguje
                        # Write total
                        clean_text = text.strip("=").strip()
                        self.current_sheet.merge_cells(f'A{self.current_row}:B{self.current_row}')
                        self.current_sheet.cell(self.current_row, 1, clean_text).font = Font(bold=True)
                elif text.startswith("-"):  # ignorujeme - 
                    pass
                elif any(keyword in text for keyword in {"Rozpis účtovaných poplatkov:", "Rozpis SMS / MMS:", "Rozpis volaní:", "Rozpis dát:"}): # jednotlive sekcie....
                    self.current_row += 2
                    self.current_sheet.cell(self.current_row, 1, text.strip()).font = Font(bold=True)
                    self.current_sheet.cell(self.current_row, 1).border = thin_border
                    if "Rozpis účtovaných poplatkov:" in text:
                        self.current_sheet.cell(self.current_row, 2, "DPH").font = Font(bold=True)
                        self.current_sheet.cell(self.current_row, 2).border = thin_border
                        self.current_sheet.cell(self.current_row, 3, "Suma v EUR bez DPH").font = Font(bold=True)
                        self.current_sheet.cell(self.current_row, 3).border = thin_border
                        
                    elif "Rozpis SMS / MMS:" in text:
                        self.current_sheet.cell(self.current_row, 2, "Počet kusov").font = Font(bold=True)
                        self.current_sheet.cell(self.current_row, 2).border = thin_border
                    elif "Rozpis volaní:" in text:
                        self.current_sheet.cell(self.current_row, 2, "Trvanie hovorov").font = Font(bold=True)
                        self.current_sheet.cell(self.current_row, 2).border = thin_border
                    elif "Rozpis dát:" in text:
                        self.current_sheet.cell(self.current_row, 2, "Spotreba dát v MB").font = Font(bold=True)
                        self.current_sheet.cell(self.current_row, 2).border = thin_border
                elif "účtovné" in text: # samostatne pre obdobie faktur
                    parts = text.split("obdobie")
                    parts[0] += "obdobie"
                    self.current_row += 1
                    self.current_sheet.cell(self.current_row,1,parts[0].strip())
                    self.current_sheet.cell(self.current_row, 1).border = thin_border
                    self.current_sheet.cell(self.current_row,2,parts[1].strip())
                    self.current_sheet.cell(self.current_row, 2).border = thin_border
                elif "NOVEM" in text:
                    self.current_row += 1
                    self.current_sheet.cell(self.current_row,1,text.strip())
                    self.current_sheet.cell(self.current_row, 1).border = thin_border
                else:
                    parts = text.strip().rsplit(maxsplit=1)
                    if len(parts) == 2 and parts[1] == "EUR":
                        last_value = parts[0].rsplit(maxsplit=1)
                        if "%" in parts[0]:
                            match = re.match(r'(.*?)(\d+%)(.*)$', text)
                            if match:
                                service = match.group(1).strip()
                                vat_rate = match.group(2)
                                remaining = match.group(3).strip()
                        else:
                            match = False
                            if len(last_value) == 2:
                                parts[0] = last_value[0]
                                parts[1] = f"{last_value[1]} {parts[1]}"
                    elif len(parts) == 2 and parts[1] in {"SMS", "MMS"}:
                        last_value = parts[0].rsplit(maxsplit=1)
                        if len(last_value) == 2:
                            parts[0] = last_value[0]
                            parts[1] = f"{last_value[1]} ks"
                    elif len(parts) == 2 and parts[1] in {"MB", "GB"}:
                        last_value = parts[0].rsplit(maxsplit=1) 
                        if len(last_value) == 2:
                            parts[0] = last_value[0]
                            parts[1] = f"{last_value[1]} {parts[1]}" 
                    #print(parts)
                    if match:
                        self.current_row += 1
                        x = f"{float(remaining.replace('EUR', '').replace(',', '.').strip()):.4f}".replace('.', ',')
                        self.current_sheet.cell(self.current_row,1,service)
                        self.current_sheet.cell(self.current_row, 1).border = thin_border
                        self.current_sheet.cell(self.current_row,2,vat_rate)
                        self.current_sheet.cell(self.current_row, 2).border = thin_border
                        self.current_sheet.cell(self.current_row, 3, x)
                        self.current_sheet.cell(self.current_row, 3).border = thin_border
                        cell_ref = self.current_sheet.cell(self.current_row, 3)
                        cell_ref.number_format = '0,0000'
                    else:
                        if len(parts) == 2:
                            self.current_row += 1
                            if "Rozpis" in parts[0] or "Faktúrovaná suma nad paušál bez DPH:" in parts[0]:
                                self.current_row +=1
                                self.current_sheet.cell(self.current_row, 1, parts[0].strip()).font = Font(bold=True)
                                self.current_sheet.cell(self.current_row, 1).border = thin_border
                            else:
                                self.current_sheet.cell(self.current_row, 1, parts[0].strip())
                                self.current_sheet.cell(self.current_row, 1).border = thin_border
                            if "EUR" in parts[1]:
                                x = f"{float(parts[1].replace('EUR', '').replace(',', '.').strip()):.4f}".replace('.', ',')
                                if "Faktúrovaná suma nad paušál bez DPH:" in parts[0] and self.current_company:
                                    amount_str = text.split(":")[-1].strip().replace("EUR", "").strip()
                                    amount = float(amount_str)
                                    self.company_totals[self.current_company] += amount
                                self.current_sheet.cell(self.current_row, 2, x)
                                self.current_sheet.cell(self.current_row, 2).border = thin_border
                                cell_ref = self.current_sheet.cell(self.current_row, 2)
                                cell_ref.number_format = '0,0000'
                            elif "MB" in parts[1]:
                                x = f"{float(parts[1].replace('MB', '').replace(',', '.').strip()):.4f}".replace('.', ',')
                                self.current_sheet.cell(self.current_row, 2, x)
                                self.current_sheet.cell(self.current_row, 2).border = thin_border
                                cell_ref = self.current_sheet.cell(self.current_row, 2)
                                cell_ref.number_format = '0,0000'
                            else:
                                self.current_sheet.cell(self.current_row, 2, parts[1].strip())
                                self.current_sheet.cell(self.current_row, 2).border = thin_border
                adjust_column_width(self.current_sheet) # toto bude treba doriesit
                
                # ukladame excel po kazdom zasahu, nech sa nam nic zle nestane :)
                self.current_xlsx_workbook.save(self.xlsx_filepath)
            except Exception as e:
                print(f"Error writing to Excel: {e}")  # Debug print


    
    def close(self):
        #print("Closing files...")  # Debug print
        if self.current_txt_file is not None:
            self.current_txt_file.close()
            self.current_txt_file = None
        
        if self.current_xlsx_workbook is not None and self.xlsx_filepath is not None:
            try:
                #print(f"Saving Excel file to: {self.xlsx_filepath}")  # Debug print
                self.current_xlsx_workbook.save(self.xlsx_filepath)
                self.current_xlsx_workbook.close()

                #print("Starting PDF conversion...")  # Debug print
                #excel_to_pdf(self.xlsx_filepath)
                #print("PDF conversion completed")  # Debug print
                
                self.current_xlsx_workbook = None
                
            except Exception as e:
                print(f"Error saving Excel file or converting to PDF: {e}")
    def write_company_summaries(self):
        """Write summary files for all companies."""
        for company_name, total in self.company_totals.items():
            # Get the company's output folder
            company_folder = get_output_folder(('company', company_name),base_dir="output_excel")
            summary_path = os.path.join(company_folder, 'nad_ramec.txt')
            
            try:
                with open(summary_path, 'w', encoding='utf-8') as f:
                    f.write(f"{total:.2f}")
            except Exception as e:
                print(f"Error writing summary for {company_name}: {e}")

def analyze_telekom_csv(file: str, output_manager: OutputManager, customer_data):
    try:
        df = pd.read_csv(file, encoding="utf-8", low_memory=False)
        grouped = df.groupby("Charges__BudgetCentre__ProductFamily__Charge__@Label")
        for number, data in grouped:
            number = clean_phone_number(number)
            customer_type_and_info = get_customer_name_or_company(customer_data, number)
            output_manager.start_new_file(number, customer_type_and_info)
            output_manager.write("=" * 80)
            output_manager.write(f"TELEKOM TELEFONNE CISLO {number}")
            tbiznis_rows = data[data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('T-Biznis', na=False)]
            if not tbiznis_rows.empty:
                date_start = tbiznis_rows['Charges__BudgetCentre__ProductFamily__Charge__@FromDate'].min()
                date_end = tbiznis_rows['Charges__BudgetCentre__ProductFamily__Charge__@ToDate'].max()

                if pd.notna(date_start) and pd.notna(date_end):  # Ensure dates are not NaN
                    formatted_date_start = datetime.strptime(str(date_start), "%Y-%m-%d").strftime("%d-%m-%Y")
                    formatted_date_end = datetime.strptime(str(date_end), "%Y-%m-%d").strftime("%d-%m-%Y")
                    output_manager.write(f"Za účtovné obdobie {formatted_date_start} - {formatted_date_end}")
                #output_manager.write(f"Za účtovné obdobie {date_start} - {date_end}")
            output_manager.write("=" * 80)
            # TODO: treba listovat aj poplatky, ktore maju nulovu sumu
            paid_services = data[data['Charges__BudgetCentre__ProductFamily__Charge__@Price'] > 0]
            #t_biznis_services = set()
            if not paid_services.empty:
                output_manager.write("\nRozpis účtovaných poplatkov:")
                output_manager.write("-" * 50)
                for _, row in paid_services.iterrows():
                    service = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    price = row['Charges__BudgetCentre__ProductFamily__Charge__@Price']
                    dph = row['Charges__BudgetCentre__ProductFamily__Charge__@VatRate']
                    if "T-Biznis" in service:
                        output_manager.write(f"{service:<50}")
                        #t_biznis_services.add(service)
                    else:
                        output_manager.write(f"{service:<50} {dph}% {price:>8.2f} EUR")

            output_manager.write("\nRozpis SMS / MMS:")
            output_manager.write("-" * 50)
            sms_usage = data[
                (data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'Ks') &
                (data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('SMS', na=False))]

            mms_usage = data[
                (data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'Ks') &
                (data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('MMS', na=False))
            ]
            if not sms_usage.empty:
                for _, row in sms_usage.iterrows():
                    service = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    count = int(float(row['Charges__BudgetCentre__ProductFamily__Charge__@Units']))
                    output_manager.write(f"{service:<50} {count:>8} SMS")
            else:
                output_manager.write(f"{'Odoslané SMS':<50} {0:>8} SMS")

            if not mms_usage.empty:
                for _, row in mms_usage.iterrows():
                    service = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    count = int(float(row['Charges__BudgetCentre__ProductFamily__Charge__@Units']))
                    output_manager.write(f"{service:<50} {count:>8} MMS")
            else:
                output_manager.write(f"{'Odoslané MMS':<50} {0:>8} MMS")
            
            output_manager.write("\nRozpis volaní:")
            output_manager.write("-" * 50)
            call_usage = data[data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'sekundy']
            if not call_usage.empty:
                for _, row in call_usage.iterrows():
                    service = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    duration = format_duration(row['Charges__BudgetCentre__ProductFamily__Charge__@Units'])
                    output_manager.write(f"{service:<50} {duration:>8}")
            else:
                output_manager.write(f"{'Odchádzajúce hovory':<50} {'00:00:00':>8}")

            output_manager.write("\nRozpis dát:")
            output_manager.write("-" * 50)
            data_usage = data[data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'MB']
            if not data_usage.empty:
                for _, row in data_usage.iterrows():
                    service = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    usage = format_data_usage(row['Charges__BudgetCentre__ProductFamily__Charge__@Units'])
                    output_manager.write(f"{service:<50} {usage:>8}")
            else:
                output_manager.write(f"{'Dáta v sieti':<50} {'0':>8}")
            
            tbiznis_total = data[
                data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('T-Biznis', na=False)
            ]['Charges__BudgetCentre__ProductFamily__Charge__@Price'].sum()
            
            total = paid_services['Charges__BudgetCentre__ProductFamily__Charge__@Price'].sum() - tbiznis_total
            
            output_manager.write("\n" + "=" * 80)
            output_manager.write(f"Faktúrovaná suma nad paušál bez DPH: {total:>8.2f} EUR")
            output_manager.write("=" * 80 + "\n")
            output_manager.close()
    except Exception as e:
        output_manager.write(f"Error processing Telekom CSV: {str(e)}")

def analyze_o2_csv(file: str, output_manager: OutputManager, customer_data):
    try:
        # O2 sa rozhodno neincludnut zaznam telefonneho cisla do kazdeho riadku, tak som spravil funkciu co to robi
        processed_file = preprocess_o2_csv(file)
        # teraz uz mozeme pokracovat rovnako ako telekom
        df = pd.read_csv(processed_file, encoding="utf-8", low_memory=False)
        grouped = df.groupby("Subscribers__Subscriber__MSISDN")

        for number, data in grouped:
            if not pd.isna(number):
                number = clean_phone_number(number)
                customer_type_and_info = get_customer_name_or_company(customer_data, number)
                output_manager.start_new_file(number, customer_type_and_info)
                output_manager.write("=" * 80)
                output_manager.write(f"O2 TELEFONNE CISLO {number}")
                datum_start = data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__PeriodStart'].iloc[0]
                datum_end = data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__PeriodEnd'].iloc[0]

                # Convert dates to the desired format
                if pd.notna(datum_start) and pd.notna(datum_end):  # Ensure dates are not NaN
                    formatted_datum_start = datetime.strptime(str(datum_start), "%Y-%m-%d").strftime("%d-%m-%Y")
                    formatted_datum_end = datetime.strptime(str(datum_end), "%Y-%m-%d").strftime("%d-%m-%Y")
                    output_manager.write(f'Za účtovné obdobie {formatted_datum_start} - {formatted_datum_end}')
                output_manager.write("=" * 80)

                recurring_fees = data[
                    (data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeType'] == 'recurring_arrears')
                ]
                if not recurring_fees.empty:
                    output_manager.write("\nRozpis účtovaných poplatkov:")
                    output_manager.write("-" * 50)
                    for _, row in recurring_fees.iterrows():
                        service = row['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName']
                        #price = row['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeNetAmount']
                        #if pd.notna(service) and pd.notna(price):
                        if pd.notna(service) and "VPN" not in service:
                            output_manager.write(f"{service:<50}")
                            #output_manager.write(f"{service:<50} {float(price):>8.2f} EUR")
                # TOTO NEBOL DOBRY PRISTUP
                # one_time_payments = data[
                #     (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Money') |
                #     (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName'].str.contains('Jednorazová platba', na=False))
                # ]
                # if not one_time_payments.empty:
                #     for _, row in one_time_payments.iterrows():
                #         service = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                #         try:
                #             price = float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemNetAmount'])
                #             dph = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__VAT']
                #         except (ValueError, TypeError):
                #             price = 0.0
                #         if pd.notna(service) and pd.notna(price):
                #             output_manager.write(f"{service:<50} {dph} {price:>8.2f} EUR")

                #TOTO JE LEPSIE
                charged_fees = data[(data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemNetAmount'] > 0)| (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Money')]
                if not charged_fees.empty:
                    output_manager.write("-" * 50)
                    for _, row in charged_fees.iterrows():
                        service = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                        usage = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Volume']
                        extra_charge = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemNetAmount']
                        dph = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__VAT']
                        output_manager.write(f"{service:<50} {dph} {extra_charge:>8.2f} EUR")

                output_manager.write("\nRozpis SMS / MMS:")
                output_manager.write("-" * 50)
                sms_usage = data[
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'occurrence') &
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName'].str.contains('SMS', na=False))]
                mms_usage = data[
                (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'occurrence') &
                (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName'].str.contains('MMS', na=False))]
                if not sms_usage.empty:
                    for _, row in sms_usage.iterrows():
                        service = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                        count = int(float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Amount']))
                        if pd.notna(service):
                            output_manager.write(f"{service:<50} {count:>8} SMS")
                else:
                    output_manager.write(f"{'Odoslané SMS':<50} {0:>8} SMS")
                if not mms_usage.empty:
                    for _, row in mms_usage.iterrows():
                        service = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                        count = int(float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Amount']))
                        if pd.notna(service):
                            output_manager.write(f"{service:<50} {count:>8} MMS")
                else:
                    output_manager.write(f"{'Odoslané MMS':<50} {0:>8} MMS")

                output_manager.write("\nRozpis volaní:")
                output_manager.write("-" * 50)
                call_usage = data[data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Second']
                if not call_usage.empty:
                    for _, row in call_usage.iterrows():
                        service = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                        duration = format_duration(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Volume'])
                        if pd.notna(service):
                            output_manager.write(f"{service:<50} {duration:>8}")
                else:
                    output_manager.write(f"{'Odchádzajúce hovory':<50} {'00:00:00':>8}")
                
                output_manager.write("\nRozpis dát:")
                output_manager.write("-" * 50)
                data_usage = data[data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Byte']
                if not data_usage.empty:
                    for _, row in data_usage.iterrows():
                        service = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                        usage = format_data_usage(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Volume'])
                        if pd.notna(service):
                            output_manager.write(f"{service:<50} {usage:>8}")
                else:
                    output_manager.write(f"{'Dáta v sieti':<50} {'0':>8} {'MB':>8}")
                
                total = data['Subscribers__Subscriber__SubscriberTotalNETAmount'].iloc[0]
                #debuging 
                #one_times = one_time_payments['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Amount'].sum()
                #if number == "421911553400":
                    #print(f'{one_times} {number}')
                e_net_total = data[
                    data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName'].str.contains('e-Net', na=False)
                ]['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeNetAmount'].sum() 
                vpn_total = data[
                    data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName'].str.contains('Mesačný poplatok VPN', na=False)
                ]['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeNetAmount'].sum()
                adjusted_total = float(total) - e_net_total - 0.03 # 0.03 je VPN poplatok O2
                output_manager.write("\n" + "=" * 80)
                output_manager.write(f"Faktúrovaná suma nad paušál bez DPH: {abs(adjusted_total):>8.2f} EUR")
                output_manager.write("=" * 80 + "\n")
                output_manager.close()
    except Exception as e:
        output_manager.write(f"Error processing O2 CSV: {str(e)}")



class InvoiceProcessor(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()

        self.title("Rozkúskovač faktúr")
        self.geometry("800x600")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.main_frame = ttk.Frame(self)
        self.main_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)

        self.text_widget = scrolledtext.ScrolledText(self.main_frame, wrap=tk.WORD, width=80, height=30)
        self.text_widget.grid(row=0, column=0, sticky="nsew")

        self.text_widget.drop_target_register(DND_FILES)
        self.text_widget.dnd_bind('<<Drop>>', self.process_dropped_file)

        # Create buttons frame
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.grid(row=1, column=0, pady=10)

        # Create buttons
        self.browse_button = ttk.Button(self.button_frame, text="Browse File", command=self.browse_file)
        self.browse_button.grid(row=0, column=0, padx=5)

        self.clear_button = ttk.Button(self.button_frame, text="Clear", command=self.clear_output)
        self.clear_button.grid(row=0, column=1, padx=5)

        # Create drop label
        #self.drop_label = ttk.Label(self.main_frame, text="Drag and drop CSV file here or use Browse button")
        #self.drop_label.grid(row=2, column=0, pady=5)

        self.output_manager = OutputManager(self.add_text)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.process_file(file_path)

    def clear_output(self):
        self.text_widget.delete(1.0, tk.END)

    def add_text(self, text: str):
        self.text_widget.insert(tk.END, text + "\n")
        self.text_widget.see(tk.END)

    def process_dropped_file(self, event):
        file_path = event.data
        # Remove curly braces if present (Windows can add these)
        file_path = file_path.strip('{}')
        self.process_file(file_path)

    def process_file(self, file_path):
        if not file_path.lower().endswith('.csv'):
            messagebox.showerror("Error", "Please select a CSV file")
            return

        self.clear_output()
        self.add_text(f"Processing file: {file_path}")

        def process_file_thread():
            customer_data = load_customer_data("zoznam_mobilky.csv")
            try:
                df = pd.read_csv(file_path, nrows=1)
                if "Charges__BudgetCentre__ProductFamily__Charge__@Label" in df.columns:
                    self.add_text("Detected Telekom invoice format")
                    analyze_telekom_csv(file_path, self.output_manager, customer_data)
                    #get_unique_tbiznis_services(file_path)
                elif "Subscribers__Subscriber__MSISDN" in df.columns:
                    self.add_text("Detected O2 invoice format")
                    analyze_o2_csv(file_path, self.output_manager, customer_data)
                    #get_unique_o2_services(file_path)
                else:
                    self.add_text("Error: Unknown CSV format")
                self.output_manager.write_company_summaries()
            except Exception as e:
                self.add_text(f"Error processing file: {str(e)}")
            
            

        threading.Thread(target=process_file_thread, daemon=True).start()

def process_csv_file(file_path: str):
    """Process CSV file without GUI"""
    print(f"Processing file: {file_path}")
    
    def console_callback(text: str):
        print(text)
    
    output_manager = OutputManager(console_callback)
    customer_data = load_customer_data("zoznam_mobilky.csv")
    
    try:
        df = pd.read_csv(file_path, nrows=1)
        if "Charges__BudgetCentre__ProductFamily__Charge__@Label" in df.columns:
            print("Detected Telekom invoice format")
            analyze_telekom_csv(file_path, output_manager, customer_data)
        elif "Subscribers__Subscriber__MSISDN" in df.columns:
            print("Detected O2 invoice format")
            analyze_o2_csv(file_path, output_manager, customer_data)
        else:
            print("Error: Unknown CSV format")
        output_manager.write_company_summaries()
    except Exception as e:
        print(f"Error processing file: {str(e)}")

def main():
    if len(sys.argv) > 1:
        # Command-line mode
        csv_file = sys.argv[1]
        if not csv_file.lower().endswith('.csv'):
            print("Error: Please provide a CSV file")
            sys.exit(1)
        process_csv_file(csv_file)
    else:
        # GUI mode
        app = InvoiceProcessor()
        app.mainloop()

if __name__ == "__main__":
    main()


