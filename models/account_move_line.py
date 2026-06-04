# Copyright 2025 NOVEM IT
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

import base64
import io
import json
import logging
import re
from collections import Counter

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pytz

from odoo import _, api, fields, models
from odoo.tools import html_escape
from odoo.tools.float_utils import float_compare

_logger = logging.getLogger(__name__)


class AccountMoveLine(models.Model):
    _inherit = 'account.move.line'

    _CUSTOMER_OVERPAYMENT_REPORT_RECIPIENTS = (
        "tomas.juricek@novem.sk,oliver.brunovsky@novem.sk,lukas.kocman@novem.sk,obrunovsky7@gmail.com"
    )
    _CUSTOMER_OVERPAYMENT_REPORT_TZ = "Europe/Bratislava"
    _CUSTOMER_OVERPAYMENT_INVOICE_REF_RE = re.compile(r"\bFAK/\d{4}/\d{5}\b")
    _CUSTOMER_OVERPAYMENT_IBAN_RE = re.compile(r"(?<![A-Z0-9])SK\d{22}(?![A-Z0-9])")
    _CUSTOMER_OVERPAYMENT_VARIABLE_SYMBOL_RE = re.compile(r"FAK/(\d{4})/(\d+)")

    stock_move_ids = fields.One2many(
        'stock.move',
        'invoice_line_id',
        string='Stock Moves',
        copy=False,
        readonly=True,
    )

    @api.model
    def _get_customer_overpayment_report_local_now(self, now_utc=None):
        timezone = pytz.timezone(self._CUSTOMER_OVERPAYMENT_REPORT_TZ)
        now_utc = now_utc or fields.Datetime.now()
        if isinstance(now_utc, str):
            now_utc = fields.Datetime.from_string(now_utc)
        if now_utc.tzinfo:
            utc_now = now_utc.astimezone(pytz.utc)
        else:
            utc_now = pytz.utc.localize(now_utc)

        return utc_now.astimezone(timezone)

    @api.model
    def _is_customer_overpayment_report_due(self, now_utc=None):
        local_now = self._get_customer_overpayment_report_local_now(
            now_utc=now_utc
        )
        return local_now.day == 15 and local_now.hour == 10

    @api.model
    def _is_customer_overpayment_report_manual_run(self):
        return (
            self.env.context.get("lastcall") is not None
            and not self.env.context.get("cron_id")
        )

    @api.model
    def _get_customer_overpayment_report_domain(self):
        return [
            ("company_id", "=", self.env.company.id),
            ("parent_state", "=", "posted"),
            ("reconciled", "=", False),
            ("partner_id", "!=", False),
            ("account_id.account_type", "=", "asset_receivable"),
            ("amount_residual", "!=", 0),
        ]

    @api.model
    def _get_customer_overpayment_report_bank_payment_domain(self):
        return [
            ("company_id", "=", self.env.company.id),
            ("is_reconciled", "=", False),
            ("amount", ">", 0),
        ]

    @api.model
    def _get_customer_overpayment_invoiced_partner_ids(self):
        invoice_partners = self.env["account.move"].sudo().with_context(
            active_test=False
        ).search([
            ("company_id", "=", self.env.company.id),
            ("state", "=", "posted"),
            ("move_type", "=", "out_invoice"),
            ("commercial_partner_id", "!=", False),
        ]).mapped("commercial_partner_id")
        return set(invoice_partners.ids)

    @api.model
    def _get_customer_overpayment_bank_text(self, statement_line):
        values = [
            statement_line.payment_ref,
            statement_line.partner_name,
            statement_line.online_partner_information,
        ]
        details = statement_line.transaction_details
        if isinstance(details, (dict, list)):
            values.append(json.dumps(details, ensure_ascii=False, sort_keys=True))
        elif details:
            values.append(str(details))
        return " ".join(value for value in values if value)

    @api.model
    def _get_customer_overpayment_invoice_refs(self, text):
        return set(self._CUSTOMER_OVERPAYMENT_INVOICE_REF_RE.findall(text or ""))

    @api.model
    def _get_customer_overpayment_ibans(self, text):
        normalized_text = re.sub(r"\s+", "", text or "").upper()
        return set(self._CUSTOMER_OVERPAYMENT_IBAN_RE.findall(normalized_text))

    @api.model
    def _match_customer_overpayment_partner_by_invoice_refs(
        self,
        invoice_refs,
    ):
        if not invoice_refs:
            return False, False, False
        invoice_moves = self.env["account.move"].sudo().with_context(
            active_test=False
        ).search([
            ("company_id", "=", self.env.company.id),
            ("state", "=", "posted"),
            ("move_type", "=", "out_invoice"),
            ("commercial_partner_id", "!=", False),
            "|",
            "|",
            ("name", "in", list(invoice_refs)),
            ("payment_reference", "in", list(invoice_refs)),
            ("ref", "in", list(invoice_refs)),
        ])
        partners = invoice_moves.mapped("commercial_partner_id")
        if len(partners) == 1:
            return partners[0], _("Referencia faktúry"), ", ".join(sorted(invoice_refs))
        if len(partners) > 1:
            return (
                False,
                _("Nejednoznačná referencia faktúry"),
                ", ".join(sorted(invoice_refs)),
            )
        return False, False, False

    @api.model
    def _match_customer_overpayment_partner_by_registered_iban(
        self,
        ibans,
        invoiced_partner_ids,
    ):
        if not ibans:
            return False, False, False
        bank_accounts = self.env["res.partner.bank"].sudo().with_context(
            active_test=False
        ).search([
            ("active", "=", True),
            ("sanitized_acc_number", "in", list(ibans)),
        ])
        partners = bank_accounts.mapped(
            "partner_id.commercial_partner_id"
        ).filtered(lambda partner: partner.id in invoiced_partner_ids)
        if len(partners) == 1:
            return partners[0], _("Registrovaný IBAN"), ", ".join(sorted(ibans))
        if len(partners) > 1:
            return (
                False,
                _("Nejednoznačný registrovaný IBAN"),
                ", ".join(sorted(ibans)),
            )
        return False, False, False

    @api.model
    def _search_customer_overpayment_statement_lines_by_token(self, token):
        like_token = "%%%s%%" % token
        self.env.cr.execute(
            """
            SELECT id
              FROM account_bank_statement_line
             WHERE company_id = %s
               AND is_reconciled = TRUE
               AND (
                    REGEXP_REPLACE(COALESCE(payment_ref, ''), '\\s+', '', 'g') ILIKE %s
                 OR REGEXP_REPLACE(COALESCE(partner_name, ''), '\\s+', '', 'g') ILIKE %s
                 OR REGEXP_REPLACE(COALESCE(online_partner_information, ''), '\\s+', '', 'g') ILIKE %s
                 OR REGEXP_REPLACE(COALESCE(transaction_details::text, ''), '\\s+', '', 'g') ILIKE %s
               )
            """,
            (
                self.env.company.id,
                like_token,
                like_token,
                like_token,
                like_token,
            ),
        )
        return self.env["account.bank.statement.line"].sudo().browse(
            [row[0] for row in self.env.cr.fetchall()]
        )

    @api.model
    def _match_customer_overpayment_partner_by_historical_iban(
        self,
        ibans,
        invoiced_partner_ids,
    ):
        if not ibans:
            return False, False, False

        matched_partners = self.env["res.partner"].sudo()
        for iban in ibans:
            history_lines = self._search_customer_overpayment_statement_lines_by_token(
                iban
            )
            for statement_line in history_lines:
                matched_partners |= statement_line.partner_id.commercial_partner_id
                matched_partners |= (
                    statement_line.move_id.partner_bank_id.partner_id
                    .commercial_partner_id
                )
                matched_partners |= statement_line.move_id.line_ids.filtered(
                    lambda line: (
                        line.partner_id
                        and line.account_id.account_type == "asset_receivable"
                    )
                ).mapped("partner_id.commercial_partner_id")

        matched_partners = matched_partners.filtered(
            lambda partner: partner.id in invoiced_partner_ids
        )
        if len(matched_partners) == 1:
            return (
                matched_partners[0],
                _("Historické párovanie IBAN"),
                ", ".join(sorted(ibans)),
            )
        if len(matched_partners) > 1:
            return (
                False,
                _("Nejednoznačné historické párovanie IBAN"),
                ", ".join(sorted(ibans)),
            )
        return False, False, False

    @api.model
    def _match_customer_overpayment_bank_payment_partner(
        self,
        statement_line,
        invoiced_partner_ids,
    ):
        statement_partner = statement_line.partner_id.commercial_partner_id
        if statement_partner and statement_partner.id in invoiced_partner_ids:
            return (
                statement_partner,
                _("Partner na bankovom riadku"),
                statement_partner.display_name,
            )

        text = self._get_customer_overpayment_bank_text(statement_line)
        invoice_refs = self._get_customer_overpayment_invoice_refs(text)
        partner, method, token = (
            self._match_customer_overpayment_partner_by_invoice_refs(invoice_refs)
        )
        if partner or method:
            return partner, method, token

        ibans = self._get_customer_overpayment_ibans(text)
        partner, method, token = (
            self._match_customer_overpayment_partner_by_registered_iban(
                ibans,
                invoiced_partner_ids,
            )
        )
        if partner or method:
            return partner, method, token

        return self._match_customer_overpayment_partner_by_historical_iban(
            ibans,
            invoiced_partner_ids,
        )

    @api.model
    def _customer_overpayment_bank_line_has_reported_receivable_credit(
        self,
        statement_line,
    ):
        currency = self.env.company.currency_id
        return any(
            line.company_id == self.env.company
            and line.parent_state == "posted"
            and not line.reconciled
            and line.partner_id
            and line.account_id.account_type == "asset_receivable"
            and line.amount_residual < 0
            and not currency.is_zero(line.amount_residual)
            for line in statement_line.move_id.line_ids
        )

    @api.model
    def _get_customer_overpayment_bank_payments_by_partner(
        self,
        invoiced_partner_ids,
    ):
        if not invoiced_partner_ids:
            return {}

        currency = self.env.company.currency_id
        bank_payments_by_partner = {}
        statement_lines = self.env["account.bank.statement.line"].sudo().with_context(
            active_test=False
        ).search(
            self._get_customer_overpayment_report_bank_payment_domain(),
            order="date asc, id asc",
        )
        for statement_line in statement_lines:
            if (
                currency.is_zero(statement_line.amount)
                or self._customer_overpayment_bank_line_has_reported_receivable_credit(
                    statement_line
                )
            ):
                continue

            partner, match_method, matched_token = (
                self._match_customer_overpayment_bank_payment_partner(
                    statement_line,
                    invoiced_partner_ids,
                )
            )
            if not partner:
                if match_method:
                    _logger.info(
                        "Customer overpayment bank payment %s skipped: %s.",
                        statement_line.id,
                        match_method,
                    )
                continue

            bank_text = self._get_customer_overpayment_bank_text(statement_line)
            invoice_refs = self._get_customer_overpayment_invoice_refs(bank_text)
            bank_payments_by_partner.setdefault(partner.id, []).append({
                "statement_line": statement_line,
                "amount": currency.round(statement_line.amount),
                "match_method": match_method,
                "matched_token": matched_token,
                "payment_ref": statement_line.payment_ref or "",
                "ibans": sorted(
                    self._get_customer_overpayment_ibans(bank_text)
                ),
                "refund_note": self._get_customer_overpayment_refund_note(
                    partner,
                    statement_line,
                    invoice_refs=invoice_refs,
                ),
            })
        return bank_payments_by_partner

    @api.model
    def _get_customer_overpayment_variable_symbol(self, move):
        match = self._CUSTOMER_OVERPAYMENT_VARIABLE_SYMBOL_RE.match(move.name or "")
        if not match:
            return ""
        return "%s%s" % (match.group(1), match.group(2))

    @api.model
    def _get_customer_overpayment_refunded_invoices(self, invoices, amount):
        currency = self.env.company.currency_id
        amount_matched_invoices = invoices.filtered(
            lambda move: float_compare(
                move.amount_total,
                amount,
                precision_rounding=currency.rounding,
            ) == 0
        )
        if not amount_matched_invoices:
            return self.env["account.move"]

        refunded_invoice_ids = set(
            self.env["account.move"].sudo().with_context(active_test=False).search([
                ("company_id", "=", self.env.company.id),
                ("state", "=", "posted"),
                ("move_type", "=", "out_refund"),
                ("reversed_entry_id", "in", amount_matched_invoices.ids),
            ]).mapped("reversed_entry_id").ids
        )
        return amount_matched_invoices.filtered(
            lambda move: move.id in refunded_invoice_ids
        )

    @api.model
    def _get_customer_overpayment_refund_note(
        self,
        partner,
        statement_line,
        invoice_refs=False,
    ):
        statement_date = statement_line.date or fields.Date.context_today(self)
        invoice_date_from = statement_date - relativedelta(months=1)
        invoice_refs = set(invoice_refs or [])
        invoice_model = self.env["account.move"].sudo().with_context(
            active_test=False
        )
        invoice_domain = [
            ("company_id", "=", self.env.company.id),
            ("state", "=", "posted"),
            ("move_type", "=", "out_invoice"),
            ("commercial_partner_id", "=", partner.id),
            ("invoice_date", ">=", invoice_date_from),
            ("invoice_date", "<=", statement_date),
        ]
        if invoice_refs:
            invoices = invoice_model.search(
                invoice_domain
                + [
                    "|",
                    "|",
                    ("name", "in", list(invoice_refs)),
                    ("payment_reference", "in", list(invoice_refs)),
                    ("ref", "in", list(invoice_refs)),
                ],
                order="invoice_date desc, id desc",
            )
            refunded_invoices = self._get_customer_overpayment_refunded_invoices(
                invoices,
                statement_line.amount,
            )
            if not refunded_invoices:
                return {"text": "", "variable_symbol": ""}
        else:
            invoices = invoice_model.search(
                invoice_domain,
                order="invoice_date desc, id desc",
            )
            refunded_invoices = self._get_customer_overpayment_refunded_invoices(
                invoices,
                statement_line.amount,
            )
            if not refunded_invoices:
                return {"text": "", "variable_symbol": ""}

        invoice = refunded_invoices[0]
        return {
            "text": _("Vrátenie mylnej platby k %s") % invoice.name,
            "variable_symbol": self._get_customer_overpayment_variable_symbol(
                invoice
            ),
        }

    @api.model
    def _get_customer_overpayment_most_used_iban(self, partner, bank_payments):
        iban_counter = Counter()
        for payment in bank_payments:
            iban_counter.update(payment.get("ibans") or [])

        partner_bank_accounts = self.env["res.partner.bank"].sudo().with_context(
            active_test=False
        ).search([
            ("active", "=", True),
            ("partner_id", "child_of", partner.id),
            ("sanitized_acc_number", "!=", False),
        ])
        iban_counter.update(partner_bank_accounts.mapped("sanitized_acc_number"))

        if not iban_counter:
            return ""
        return sorted(
            iban_counter.items(),
            key=lambda item: (-item[1], item[0]),
        )[0][0]

    @api.model
    def _get_customer_overpayment_report_data(self):
        line_model = self.sudo().with_context(active_test=False)
        currency = self.env.company.currency_id
        base_domain = self._get_customer_overpayment_report_domain()
        trigger_lines = line_model.search(base_domain + [("amount_residual", "<", 0)])
        receivable_partners = trigger_lines.mapped("partner_id.commercial_partner_id")
        invoiced_partner_ids = self._get_customer_overpayment_invoiced_partner_ids()
        bank_payments_by_partner = (
            self._get_customer_overpayment_bank_payments_by_partner(
                invoiced_partner_ids
            )
        )
        partner_ids = set(receivable_partners.ids) | set(bank_payments_by_partner)
        if not partner_ids:
            return []

        partners = self.env["res.partner"].sudo().with_context(
            active_test=False
        ).browse(list(partner_ids)).exists()
        open_lines = line_model.search(
            base_domain + [("partner_id", "child_of", partners.ids)],
            order="date asc, id asc",
        )

        report_data = []
        for partner in partners.sorted("display_name"):
            partner_lines = open_lines.filtered(
                lambda line: line.partner_id.commercial_partner_id == partner
            )
            negative_total = sum(
                line.amount_residual
                for line in partner_lines
                if line.amount_residual < 0
            )
            receivable_overpaid_amount = 0
            if negative_total < 0 and not currency.is_zero(negative_total):
                receivable_overpaid_amount = currency.round(abs(negative_total))
            bank_payments = bank_payments_by_partner.get(partner.id, [])
            bank_overpaid_amount = currency.round(
                sum(payment["amount"] for payment in bank_payments)
            )
            overpaid_amount = currency.round(
                receivable_overpaid_amount + bank_overpaid_amount
            )
            if currency.is_zero(overpaid_amount):
                continue
            report_data.append({
                "partner": partner,
                "lines": partner_lines,
                "bank_payments": bank_payments,
                "most_used_iban": self._get_customer_overpayment_most_used_iban(
                    partner,
                    bank_payments,
                ),
                "overpaid_amount": overpaid_amount,
                "receivable_overpaid_amount": receivable_overpaid_amount,
                "bank_overpaid_amount": bank_overpaid_amount,
                "total_balance": currency.round(
                    sum(partner_lines.mapped("amount_residual"))
                ),
                "line_count": len(partner_lines),
            })

        return sorted(
            report_data,
            key=lambda data: (-data["overpaid_amount"], data["partner"].display_name),
        )

    @api.model
    def _format_customer_overpayment_amount(self, amount):
        return self.env.company.currency_id.format(amount)

    @api.model
    def _format_customer_overpayment_date(self, date_value):
        return fields.Date.to_string(date_value) if date_value else ""

    @api.model
    def _get_customer_overpayment_document_type(self, move):
        return {
            "out_invoice": _("Odberateľská faktúra"),
            "out_refund": _("Odberateľský dobropis"),
            "entry": _("Účtovný zápis"),
            "out_receipt": _("Predajný doklad"),
        }.get(move.move_type, move.move_type or "")

    @api.model
    def _get_customer_overpayment_partner_url(self, partner):
        base_url = self.env["ir.config_parameter"].sudo().get_param("web.base.url")
        if not base_url:
            return ""
        return "%s/web#id=%s&model=res.partner&view_type=form" % (
            base_url.rstrip("/"),
            partner.id,
        )

    @api.model
    def _get_customer_overpayment_sheet_name(self, partner, used_names):
        base_name = re.sub(r"[\[\]\:\*\?\/\\]", " ", partner.display_name or "Zákazník")
        base_name = re.sub(r"\s+", " ", base_name).strip() or "Zákazník"
        base_name = base_name[:31]
        sheet_name = base_name
        counter = 2
        while sheet_name in used_names:
            suffix = " %s" % counter
            sheet_name = "%s%s" % (base_name[:31 - len(suffix)], suffix)
            counter += 1
        used_names.add(sheet_name)
        return sheet_name

    @api.model
    def _generate_customer_overpayment_report_xlsx(self, report_data, report_date):
        workbook = Workbook()
        currency = self.env.company.currency_id
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        alternate_fill = PatternFill("solid", fgColor="F5F5F5")
        negative_font = Font(color="B42318", bold=True)
        header_font = Font(bold=True)
        date_value = (
            fields.Date.from_string(report_date)
            if isinstance(report_date, str)
            else report_date
        )

        summary_sheet = workbook.active
        summary_sheet.title = "Súhrn"
        summary_rows = [
            ["Prehľad preplatkov zákazníkov"],
            ["Dátum kontroly", date_value],
            ["Počet zákazníkov", len(report_data)],
            [
                "Preplatky spolu",
                sum(data["overpaid_amount"] for data in report_data),
            ],
            [],
            ["Zákazník", "Preplatok spolu", "Celkový otvorený zostatok"],
        ]
        for row in summary_rows:
            summary_sheet.append(row)
        summary_sheet["A1"].font = Font(bold=True, size=14)
        summary_sheet["B2"].number_format = "DD.MM.YYYY"
        for cell in summary_sheet[6]:
            cell.font = header_font
            cell.fill = header_fill
        for index, data in enumerate(report_data, start=7):
            summary_sheet.append([
                data["partner"].display_name,
                data["overpaid_amount"],
                data["total_balance"],
            ])
            if index % 2 == 0:
                for cell in summary_sheet[index]:
                    cell.fill = alternate_fill
            summary_sheet.cell(index, 2).font = negative_font
        for column in range(1, 4):
            summary_sheet.column_dimensions[get_column_letter(column)].width = 28
        for row in summary_sheet.iter_rows(min_row=7, min_col=2, max_col=3):
            for cell in row:
                cell.number_format = '#,##0.00 "%s"' % currency.symbol
                cell.alignment = Alignment(horizontal="right")

        used_names = {"Súhrn"}
        detail_headers = [
            "Zdroj",
            "Dátum",
            "Splatnosť",
            "Účtovný zápis",
            "Typ",
            "Popis / referencia",
            "Suma",
            "Metóda párovania",
            "Spárovaný token",
            "IBAN",
            "Poznámka",
            "VS",
        ]
        for data in report_data:
            sheet = workbook.create_sheet(
                self._get_customer_overpayment_sheet_name(
                    data["partner"],
                    used_names,
                )
            )
            sheet.append([data["partner"].display_name])
            sheet.append(["Preplatok spolu", data["overpaid_amount"]])
            sheet.append(["Celkový otvorený zostatok", data["total_balance"]])
            sheet.append(["Preplatky v saldokonte", data["receivable_overpaid_amount"]])
            sheet.append(["Nespárované bankové platby", data["bank_overpaid_amount"]])
            sheet.append(["Najčastejšie používaný IBAN", data["most_used_iban"]])
            sheet.append([])
            sheet.append(detail_headers)
            sheet["A1"].font = Font(bold=True, size=14)
            for cell in sheet[8]:
                cell.font = header_font
                cell.fill = header_fill
            detail_rows = []
            for line in data["lines"]:
                label = line.name or line.ref or line.move_id.ref or ""
                detail_rows.append({
                    "date": line.date,
                    "id": line.id,
                    "amount": line.amount_residual,
                    "values": [
                        "Saldokonto zákazníka",
                        line.date,
                        line.date_maturity,
                        line.move_id.name or "",
                        self._get_customer_overpayment_document_type(line.move_id),
                        label,
                        line.amount_residual,
                        "",
                        "",
                        data["most_used_iban"],
                        "",
                        "",
                    ],
                })

            for payment in data["bank_payments"]:
                statement_line = payment["statement_line"]
                refund_note = payment.get("refund_note") or {}
                detail_rows.append({
                    "date": statement_line.date,
                    "id": statement_line.id,
                    "amount": -payment["amount"],
                    "values": [
                        "Nespárovaná platba",
                        statement_line.date,
                        "",
                        statement_line.move_id.name or "",
                        "Riadok bankového výpisu",
                        payment["payment_ref"],
                        -payment["amount"],
                        payment["match_method"],
                        payment["matched_token"],
                        data["most_used_iban"],
                        refund_note.get("text", ""),
                        refund_note.get("variable_symbol", ""),
                    ],
                })

            detail_rows = sorted(
                detail_rows,
                key=lambda row: (
                    row["amount"] < 0,
                    row["date"] or fields.Date.from_string("1900-01-01"),
                    row["id"],
                ),
            )
            row_index = 9
            for detail_row in detail_rows:
                sheet.append(detail_row["values"])
                if row_index % 2 == 0:
                    for cell in sheet[row_index]:
                        cell.fill = alternate_fill
                for date_cell in (sheet.cell(row_index, 2), sheet.cell(row_index, 3)):
                    date_cell.number_format = "DD.MM.YYYY"
                amount_cell = sheet.cell(row_index, 7)
                amount_cell.number_format = '#,##0.00 "%s"' % currency.symbol
                amount_cell.alignment = Alignment(horizontal="right")
                if detail_row["amount"] < 0:
                    amount_cell.font = negative_font
                row_index += 1

            sheet.column_dimensions["A"].width = 24
            sheet.column_dimensions["B"].width = 14
            sheet.column_dimensions["C"].width = 14
            sheet.column_dimensions["D"].width = 22
            sheet.column_dimensions["E"].width = 22
            sheet.column_dimensions["F"].width = 60
            sheet.column_dimensions["G"].width = 18
            sheet.column_dimensions["H"].width = 24
            sheet.column_dimensions["I"].width = 36
            sheet.column_dimensions["J"].width = 28
            sheet.column_dimensions["K"].width = 42
            sheet.column_dimensions["L"].width = 14
            for row in sheet.iter_rows(min_row=2, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = '#,##0.00 "%s"' % currency.symbol
                    cell.alignment = Alignment(horizontal="right")
            sheet.freeze_panes = "A9"

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @api.model
    def _render_customer_overpayment_report_email(self, report_data, report_date):
        report_date_display = fields.Date.to_string(report_date)
        total_overpaid = sum(data["overpaid_amount"] for data in report_data)

        table_style = (
            "border-collapse:collapse;width:100%;margin:12px 0 20px 0;"
            "font-size:13px;"
        )
        th_style = (
            "border:1px solid #d8dee4;background:#f6f8fa;padding:6px 8px;"
            "text-align:left;"
        )
        td_style = "border:1px solid #d8dee4;padding:6px 8px;vertical-align:top;"
        alternate_td_style = td_style + "background:#f7f7f7;"
        amount_style = td_style + "text-align:right;white-space:nowrap;"
        alternate_amount_style = (
            alternate_td_style + "text-align:right;white-space:nowrap;"
        )
        negative_style = amount_style + "color:#b42318;font-weight:600;"
        alternate_negative_style = (
            alternate_amount_style + "color:#b42318;font-weight:600;"
        )

        html = [
            "<div>",
            "<h2>Prehľad preplatkov zákazníkov</h2>",
            "<p>",
            "Kontrola našla zákazníkov s preplatkom v saldokonte alebo "
            "s konzervatívne spárovanou nespárovanou bankovou platbou.",
            "</p>",
            "<p>",
            "<strong>Dátum kontroly:</strong> %s<br/>" % html_escape(report_date_display),
            "<strong>Počet zákazníkov:</strong> %s<br/>" % len(report_data),
            "<strong>Preplatky spolu:</strong> %s"
            % html_escape(self._format_customer_overpayment_amount(total_overpaid)),
            "</p>",
            '<table style="%s">' % table_style,
            "<thead><tr>",
            '<th style="%s">Zákazník</th>' % th_style,
            '<th style="%s">Preplatok spolu</th>' % th_style,
            '<th style="%s">Celkový otvorený zostatok</th>' % th_style,
            "</tr></thead><tbody>",
        ]

        for index, data in enumerate(report_data):
            row_td_style = alternate_td_style if index % 2 else td_style
            row_amount_style = alternate_amount_style if index % 2 else amount_style
            row_negative_style = (
                alternate_negative_style if index % 2 else negative_style
            )
            partner = data["partner"]
            partner_name = html_escape(partner.display_name)
            partner_url = self._get_customer_overpayment_partner_url(partner)
            if partner_url:
                partner_name = (
                    '<a href="%s" target="_blank" rel="noopener noreferrer">%s</a>'
                    % (html_escape(partner_url), partner_name)
                )
            html.extend([
                "<tr>",
                '<td style="%s">%s</td>'
                % (row_td_style, partner_name),
                '<td style="%s">%s</td>'
                % (
                    row_negative_style,
                    html_escape(
                        self._format_customer_overpayment_amount(
                            data["overpaid_amount"]
                        )
                    ),
                ),
                '<td style="%s">%s</td>'
                % (
                    row_amount_style,
                    html_escape(
                        self._format_customer_overpayment_amount(
                            data["total_balance"]
                        )
                    ),
                ),
                "</tr>",
            ])
        html.extend(["</tbody></table>"])
        html.append(
            "<p>Detailné otvorené riadky a spárované nespárované bankové "
            "platby sú v priloženom Excel súbore, každý zákazník má "
            "vlastný hárok.</p>"
        )

        html.append("</div>")
        return "".join(html)

    @api.model
    def _send_customer_overpayment_report_email(self, report_data, report_date=None):
        report_date = report_date or fields.Date.context_today(self)
        body_html = self._render_customer_overpayment_report_email(
            report_data,
            report_date,
        )
        filename = "prehlad_preplatkov_zakaznikov_%s.xlsx" % fields.Date.to_string(
            report_date
        )
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "datas": base64.b64encode(
                self._generate_customer_overpayment_report_xlsx(
                    report_data,
                    report_date,
                )
            ),
            "mimetype": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            "type": "binary",
        })
        mail_values = {
            "subject": _("Prehľad preplatkov zákazníkov - %s")
            % fields.Date.to_string(report_date),
            "email_from": self.env.company.email or self.env.user.email,
            "email_to": self._CUSTOMER_OVERPAYMENT_REPORT_RECIPIENTS,
            "body_html": body_html,
            "attachment_ids": [(4, attachment.id)],
            "auto_delete": False,
        }
        return (
            self.env["mail.mail"]
            .sudo()
            .with_context(
                mail_notify_force_send=False,
                mail_auto_subscribe_no_notify=True,
                tracking_disable=True,
                mail_create_nolog=True,
            )
            .create(mail_values)
            .send()
        )

    @api.model
    def _check_and_send_customer_overpayment_report(
        self,
        now_utc=None,
        force_send=False,
    ):
        manual_run = self._is_customer_overpayment_report_manual_run()
        if (
            not force_send
            and not manual_run
            and not self._is_customer_overpayment_report_due(now_utc=now_utc)
        ):
            _logger.info(
                "Customer overpayment report skipped outside the 15th day "
                "10:00 Bratislava hour."
            )
            return False
        if manual_run and not force_send:
            _logger.info(
                "Customer overpayment report manually triggered; "
                "bypassing the 15th day 10:00 Bratislava schedule gate."
            )

        report_data = self._get_customer_overpayment_report_data()
        if not report_data:
            _logger.info("Customer overpayment report found no overpaid customers.")
            return False

        local_now = self._get_customer_overpayment_report_local_now(
            now_utc=now_utc
        )
        self._send_customer_overpayment_report_email(
            report_data,
            report_date=local_now.date(),
        )
        _logger.info(
            "Customer overpayment report sent for %s customer(s).",
            len(report_data),
        )
        return True

    @api.model
    def cron_send_customer_overpayment_report(self, force_send=False):
        return self._check_and_send_customer_overpayment_report(
            force_send=force_send,
        )
