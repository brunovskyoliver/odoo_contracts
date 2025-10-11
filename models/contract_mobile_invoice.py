# Copyright 2025
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

from odoo import api, fields, models, _
from odoo.exceptions import UserError
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import pandas as pd
import re
import io
import csv
import base64
import logging

_logger = logging.getLogger(__name__)

def insert_image(sheet, image_path, cell):
    """Insert an image into the specified cell of the sheet."""
    try:
        # Load the image
        img = Image(image_path)
        
        # Adjust the size of the image if needed
        img.width, img.height = 200, 100
        
        # Extract column letter and row number from cell reference
        col_letter = ''.join(filter(str.isalpha, cell))
        row_num = int(''.join(filter(str.isdigit, cell)))
        
        # Get column width in pixels (1 unit = 7 pixels)
        col_width = sheet.column_dimensions[col_letter].width
        col_width_px = col_width * 7
        
        # Calculate offsets to center the image
        x_offset = (col_width_px - img.width) / 2
        
        # Adjust row height to fit the image
        row_height = img.height * 0.75  # Convert to points (Excel units)
        sheet.row_dimensions[row_num].height = row_height
        
        # Calculate y_offset to center vertically
        y_offset = (row_height - img.height * 0.75) / 2
        
        # Add the image with calculated offsets
        img.anchor = cell
        img.left = x_offset
        img.top = y_offset
        
        sheet.add_image(img)
        
    except Exception as e:
        _logger.error(f"Error inserting image: {e}")

def format_duration(seconds):
    """Format seconds into HH:MM:SS format."""
    try:
        seconds = float(seconds)
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    except ValueError:
        return "00:00:00"

def format_data_usage(bytes_or_mb):
    """Format bytes or MB into appropriate unit (MB/GB)."""
    try:
        # Convert the input to float and handle MB conversion
        value = float(bytes_or_mb)
        mb = value / (1024 * 1024) if value > 1024 * 1024 else value
        
        # Format as GB if over 1024 MB, otherwise as MB
        if mb > 1024:
            return f"{mb/1024:.2f} GB"
        else:
            return f"{mb:.2f} MB"
    except ValueError:
        return "0 MB"

def get_output_folder(customer_info: tuple, base_dir: str = "output") -> str:
    """Determine the appropriate output folder based on customer info."""
    customer_type, name = customer_info
    
    # Create base output directory if it doesn't exist
    os.makedirs(base_dir, exist_ok=True)
    
    # Create companies and individuals directories
    companies_dir = os.path.join(base_dir, "companies")
    individuals_dir = os.path.join(base_dir, "individuals")
    os.makedirs(companies_dir, exist_ok=True)
    os.makedirs(individuals_dir, exist_ok=True)
    
    if customer_type == 'company':
        company_dir = os.path.join(companies_dir, name)
        os.makedirs(company_dir, exist_ok=True)
        return company_dir
    else:  # 'individual' or 'unknown'
        return individuals_dir


class ContractMobileInvoice(models.Model):
    _name = "contract.mobile.invoice"
    _description = "Mobile Service Invoice"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "date desc"

    name = fields.Char(string="Reference", required=True, tracking=True)
    date = fields.Date(string="Invoice Date", required=True, tracking=True)
    operator = fields.Selection(
        selection=[
            ('telekom', 'Telekom'),
            ('o2', 'O2'),
        ],
        string="Operator",
        required=True,
        tracking=True,
    )
    state = fields.Selection(
        selection=[
            ('draft', 'Draft'),
            ('processed', 'Processed'),
            ('done', 'Done'),
        ],
        string="State",
        default='draft',
        tracking=True,
    )
    invoice_line_ids = fields.One2many(
        comodel_name="contract.mobile.invoice.line",
        inverse_name="invoice_id",
        string="Invoice Lines",
    )
    line_count = fields.Integer(
        string="Number of Lines",
        compute="_compute_line_count",
        store=True,
    )
    company_id = fields.Many2one(
        comodel_name="res.company",
        string="Company",
        default=lambda self: self.env.company,
    )
    csv_file = fields.Binary(string="CSV File", attachment=True)
    csv_filename = fields.Char(string="CSV File Name")
    notes = fields.Text(string="Notes")
    
    @api.depends('invoice_line_ids')
    def _compute_line_count(self):
        for record in self:
            record.line_count = len(record.invoice_line_ids)
    
    def action_process_invoice(self):
        """Process the CSV file and create invoice lines"""
        self.ensure_one()
        
        if self.csv_file and self.operator:
            # Delete existing invoice lines
            self.invoice_line_ids.unlink()
            
            try:
                # Decode the CSV file
                csv_content = base64.b64decode(self.csv_file)
                
                if self.operator == 'telekom':
                    lines = self._process_telekom_csv(csv_content)
                elif self.operator == 'o2':
                    lines = self._process_o2_csv(csv_content)
                else:
                    raise UserError(_("Unsupported operator: %s") % self.operator)
                    
                # Create invoice lines
                invoice_lines = []
                for line in lines:
                    invoice_lines.append((0, 0, line))
                    
                self.write({
                    'invoice_line_ids': invoice_lines,
                    'state': 'processed',
                })
                
                return True
                
            except Exception as e:
                raise UserError(_("Error processing CSV file: %s") % str(e))
        
        return False
        
    def _process_telekom_csv(self, csv_content):
        """Process Telekom CSV file"""
        result = []
        
        try:
            # Read the CSV file into a pandas DataFrame
            df_raw = pd.read_csv(io.BytesIO(csv_content), encoding='utf-8', low_memory=False)
            
            # Group by phone number (Label column)
            grouped = df_raw.groupby("Charges__BudgetCentre__ProductFamily__Charge__@Label")
            
            for phone_number, data in grouped:
                if pd.isna(phone_number):
                    continue
                # Clean phone number to keep only numeric characters
                phone_number = self._clean_phone_number(phone_number)
                # Skip non-Slovak numbers
                if not phone_number.startswith('421'):
                    continue
                
                # Process T-Biznis services (basic plans)
                tbiznis_services = data[data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('T-Biznis', na=False)]
                
                for _, row in tbiznis_services.iterrows():
                    service_name = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    price = self._safe_convert_to_float(row['Charges__BudgetCentre__ProductFamily__Charge__@Price'])
                    vat_rate = str(row['Charges__BudgetCentre__ProductFamily__Charge__@VatRate']) + "%"
                    
                    if pd.notna(service_name):
                        # Map T-Biznis service names to NOVEM names
                        service_name = format_plan_name(service_name)
                        
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'basic',
                            'amount': price,
                            'total': price,
                            'is_excess_usage': False,
                            'vat': vat_rate,
                        })
                
                # Process paid services (excluding T-Biznis)
                paid_services = data[
                    (data['Charges__BudgetCentre__ProductFamily__Charge__@Price'] > 0) &
                    (~data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('T-Biznis', na=False))
                ]
                
                for _, row in paid_services.iterrows():
                    service_name = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    price = self._safe_convert_to_float(row['Charges__BudgetCentre__ProductFamily__Charge__@Price'])
                    vat_rate = str(row['Charges__BudgetCentre__ProductFamily__Charge__@VatRate']) + "%"

                    if pd.notna(service_name):
                        service_type = 'other'
                        if 'data' in service_name.lower():
                            service_type = 'data'
                        elif 'voice' in service_name.lower() or 'call' in service_name.lower():
                            service_type = 'voice'
                        elif 'sms' in service_name.lower():
                            service_type = 'sms'
                        elif 'mms' in service_name.lower():
                            service_type = 'mms'
                        elif 'roaming' in service_name.lower():
                            service_type = 'roaming'
                        
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': service_type,
                            'amount': price,
                            'total': price,
                            'is_excess_usage': True,
                            'vat': vat_rate,
                        })
                
                # Process SMS usage
                sms_usage = data[
                    (data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'Ks') &
                    (data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('SMS', na=False))
                ]
                
                for _, row in sms_usage.iterrows():
                    service_name = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    count = int(float(row['Charges__BudgetCentre__ProductFamily__Charge__@Units']))
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'sms',
                            'amount': 0.0,
                            'quantity': count,
                            'unit': 'SMS',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
                
                # Process MMS usage
                mms_usage = data[
                    (data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'Ks') &
                    (data['Charges__BudgetCentre__ProductFamily__Charge__@Desc'].str.contains('MMS', na=False))
                ]
                
                for _, row in mms_usage.iterrows():
                    service_name = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    count = int(float(row['Charges__BudgetCentre__ProductFamily__Charge__@Units']))
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'mms',
                            'amount': 0.0,
                            'quantity': count,
                            'unit': 'MMS',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
                
                # Process call usage
                call_usage = data[data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'sekundy']
                for _, row in call_usage.iterrows():
                    service_name = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    duration = float(row['Charges__BudgetCentre__ProductFamily__Charge__@Units'])
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'voice',
                            'amount': 0.0,
                            'quantity': duration,
                            'unit': 'Second',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
                
                # Process data usage
                data_usage = data[data['Charges__BudgetCentre__ProductFamily__Charge__@UnitType'] == 'MB']
                for _, row in data_usage.iterrows():
                    service_name = row['Charges__BudgetCentre__ProductFamily__Charge__@Desc']
                    volume = float(row['Charges__BudgetCentre__ProductFamily__Charge__@Units'])
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'data',
                            'amount': 0.0,
                            'quantity': volume,
                            'unit': 'MB',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
            
        except Exception as e:
            _logger.error(f"Error processing Telekom CSV file: {str(e)}")
            raise
            
        return result
        
    def _process_o2_csv(self, csv_content):
        """Process O2 CSV file"""
        result = []
        
        try:
            # First, preprocess to fill in missing phone numbers
            # Read the CSV file into a pandas DataFrame
            df_raw = pd.read_csv(io.BytesIO(csv_content), encoding='utf-8', low_memory=False)
            
            # Find the MSISDN column
            msisdn_col = None
            for col in df_raw.columns:
                if 'MSISDN' in col:
                    msisdn_col = col
                    break
            if not msisdn_col:
                raise UserError(_("Could not find MSISDN column in O2 CSV file"))
            # Forward fill the phone numbers (fill missing values with the last valid value)
            df_raw[msisdn_col] = df_raw[msisdn_col].fillna(method='ffill')
            # Group by phone number
            grouped = df_raw.groupby(msisdn_col)
            for phone_number, data in grouped:
                if pd.isna(phone_number):
                    continue
                # Clean phone number to keep only numeric characters
                phone_number = self._clean_phone_number(phone_number)
                # Skip non-Slovak numbers
                if not phone_number.startswith('421'):
                    continue
                # Get the basic plan amount from SubscriberTotalNETAmount
                basic_plan_amount = self._safe_convert_to_float(data['Subscribers__Subscriber__SubscriberTotalNETAmount'].iloc[0])
                
                # Process recurring fees
                recurring_fees = data[
                    (data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeType'] == 'recurring_arrears') &
                    (~data['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName'].str.contains('VPN', na=False))
                ]
                
                for _, row in recurring_fees.iterrows():
                    service_name = row['Subscribers__Subscriber__InvoiceLines__FeeLines__Fee__FeeName']
                    service_name = handle_o2_service_name(service_name)
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'basic',
                            'amount': basic_plan_amount,  # Use the basic plan amount from SubscriberTotalNETAmount
                            'total': basic_plan_amount,
                            'is_excess_usage': False,
                        })
                
                # Process charged fees (one-time payments and extra charges)
                charged_fees = data[
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemNetAmount'] > 0) |
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Money')
                ]
                
                for _, row in charged_fees.iterrows():
                    service_name = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                    amount = self._safe_convert_to_float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemNetAmount'])
                    vat = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__VAT']
                    
                    if pd.notna(service_name):
                        service_type = 'other'
                        if 'data' in service_name.lower():
                            service_type = 'data'
                        elif 'voice' in service_name.lower() or 'call' in service_name.lower():
                            service_type = 'voice'
                        elif 'sms' in service_name.lower():
                            service_type = 'sms'
                        elif 'mms' in service_name.lower():
                            service_type = 'mms'
                        elif 'roaming' in service_name.lower():
                            service_type = 'roaming'
                        
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': service_type,
                            'amount': amount,
                            'total': amount,
                            'vat': vat,
                            'is_excess_usage': True,
                        })
                
                # Process SMS/MMS usage
                sms_usage = data[
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'occurrence') &
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName'].str.contains('SMS', na=False))
                ]
                
                mms_usage = data[
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'occurrence') &
                    (data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName'].str.contains('MMS', na=False))
                ]
                
                for _, row in sms_usage.iterrows():
                    service_name = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                    count = int(float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Amount']))
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'sms',
                            'amount': 0.0,
                            'quantity': count,
                            'unit': 'SMS',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
                
                for _, row in mms_usage.iterrows():
                    service_name = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                    count = int(float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Amount']))
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'mms',
                            'amount': 0.0,
                            'quantity': count,
                            'unit': 'MMS',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
                
                # Process call usage
                call_usage = data[data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Second']
                for _, row in call_usage.iterrows():
                    service_name = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                    duration = float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Volume'])
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'voice',
                            'amount': 0.0,
                            'quantity': duration,
                            'unit': 'Second',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
                
                # Process data usage
                data_usage = data[data['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageUOM'] == 'Byte']
                for _, row in data_usage.iterrows():
                    service_name = row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__UsageItemName']
                    volume = float(row['Subscribers__Subscriber__InvoiceLines__UsageLines__Usage__Volume'])
                    if pd.notna(service_name):
                        result.append({
                            'phone_number': phone_number,
                            'service_name': service_name,
                            'service_type': 'data',
                            'amount': 0.0,
                            'quantity': volume,
                            'unit': 'Byte',
                            'total': 0.0,
                            'is_excess_usage': False,
                        })
            
        except Exception as e:
            _logger.error(f"Error processing O2 CSV file: {str(e)}")
            raise
            
        return result
    
    # Legacy methods removed - only using implementations with parameters
    
    def action_done(self):
        """Mark invoice as done and process excess usage"""
        self.ensure_one()
        
        # First mark as done
        self.write({'state': 'done'})
        
        # Track mismatches between service names
        mismatched_services = []
        
        # Check each invoice line's service name against its mobile service
        for line in self.invoice_line_ids:
            if line.service_type == 'basic' and line.mobile_service_id:
                mobile_service = line.mobile_service_id
                if line.service_name != mobile_service.name:
                    mismatched_services.append({
                        'phone_number': line.phone_number,
                        'current_service': line.service_name,
                        'expected_service': mobile_service.name,
                        'partner': line.partner_id.name
                    })
        
        # Send email report if there are mismatches
        if mismatched_services:
            # Prepare email body
            email_body = "Report po importe mobiliek:\n\n"
            for mismatch in mismatched_services:
                email_body += f"Telefónne číslo: {mismatch['phone_number']}\n"
                email_body += f"Partner: {mismatch['partner']}\n"
                email_body += f"Súčasná služba: {mismatch['current_service']}\n"
                email_body += f"Očakávaná služba: {mismatch['expected_service']}\n"
                email_body += "-" * 50 + "\n"
            
            # Send email
            mail_values = {
                'subject': _('Report po importe mobiliek - %s') % fields.Datetime.now(),
                'email_from': self.env.company.email or self.env.user.email,
                'email_to': 'obrunovsky7@gmail.com',
                'body_html': '<pre>%s</pre>' % email_body,
                'auto_delete': False,
            }
            self.env['mail.mail'].sudo().with_context(
                mail_notify_force_send=False,
                mail_auto_subscribe_no_notify=True,
                tracking_disable=True,
                mail_create_nolog=True
            ).create(mail_values).send()
        
        # Get all invoice lines with excess usage for all partners
        excess_usage_by_partner = {}
        
        for line in self.invoice_line_ids.filtered(
            lambda l: l.is_excess_usage and l.partner_id
        ):
            if line.partner_id.id not in excess_usage_by_partner:
                excess_usage_by_partner[line.partner_id.id] = {
                    'total_23': 0.0,  # For 23% VAT
                    'total_0': 0.0,   # For 0% VAT
                    'partner': line.partner_id,
                }
            
            # Sum amounts based on VAT rate
            if line.vat == '23%':
                excess_usage_by_partner[line.partner_id.id]['total_23'] += line.total
            else:  # Assume 0% for all other cases
                excess_usage_by_partner[line.partner_id.id]['total_0'] += line.total

        _logger.info("Processing excess usage for all partners")
        
        # Process each partner's excess usage
        for partner_data in excess_usage_by_partner.values():
            # Skip if no excess usage in either VAT rate
            if partner_data['total_23'] <= 0 and partner_data['total_0'] <= 0:
                continue
                
            _logger.info(f"Found excess usage totals - 23% VAT: {partner_data['total_23']}, 0% VAT: {partner_data['total_0']} for partner {partner_data['partner'].name}")
            
            # Find the mobilky contract for this partner
            contract = self.env['contract.contract'].search([
                ('partner_id', '=', partner_data['partner'].id),
                ('x_contract_type', '=', 'Mobilky')
            ])
            if len(contract) > 1:
                for c in contract:
                    if c.nadspotreba:
                        contract = c
                        break
            if len(contract) > 1:
                raise UserError(_("Found multiple Mobilky contracts for partner {partner_data['partner'].name}"))
            if not contract:
                _logger.info(f"No active Mobilky contract found for partner {partner_data['partner'].name}")
                continue
                
            _logger.info(f"Found contract: {contract.name}")
            
            # Get products for both VAT rates
            product_23 = self.env['product.product'].search([
                ('display_name', '=', 'Vyúčtovanie paušálnych služieb a spotreby HLAS 23%')
            ], limit=1)
            
            product_0 = self.env['product.product'].search([
                ('display_name', '=', 'Vyúčtovanie paušálnych služieb a spotreby HLAS 0%')
            ], limit=1)
            
            if not (product_23 and product_0):
                _logger.error("Could not find both Vyúčtovanie products (23% and 0%)")
                continue
            
            # Handle 23% VAT line
            if partner_data['total_23'] > 0:
                existing_line_23 = contract.contract_line_ids.filtered(
                    lambda l: l.product_id == product_23
                )
                
                if existing_line_23:
                    new_total = existing_line_23.price_unit + partner_data['total_23']
                    _logger.info(f"Updating existing 23% VAT line - current: {existing_line_23.price_unit}, adding: {partner_data['total_23']}, new total: {new_total}")
                    
                    existing_line_23.with_context(skip_date_check=True).write({
                        'price_unit': new_total,  # Adjust for 23% VAT
                        'x_zlavnena_cena': new_total,  # Adjust for 23% VAT
                        'date_start': contract.recurring_next_date,
                        'recurring_next_date': contract.recurring_next_date
                    })
                else:
                    _logger.info(f"Creating new 23% VAT line with amount: {partner_data['total_23']}")
                    self.env['contract.line'].with_context(skip_date_check=True).create({
                        'contract_id': contract.id,
                        'product_id': product_23.id,
                        'name': 'Vyúčtovanie paušálnych služieb a spotreby HLAS 23%',
                        'quantity': 1,
                        'price_unit': partner_data['total_23'],  # Adjust for 23% VAT
                        'recurring_rule_type': 'monthly',
                        'recurring_interval': 1,
                        "uom_id": 1,
                        "x_zlavnena_cena": partner_data['total_23'],  # Adjust for 23% VAT
                        'date_start': contract.recurring_next_date,
                        'recurring_next_date': contract.recurring_next_date,
                        'is_auto_renew': False,
                    })
            
            # Handle 0% VAT line
            if partner_data['total_0'] > 0:
                existing_line_0 = contract.contract_line_ids.filtered(
                    lambda l: l.product_id == product_0
                )
                
                if existing_line_0:
                    new_total = existing_line_0.price_unit + partner_data['total_0']
                    _logger.info(f"Updating existing 0% VAT line - current: {existing_line_0.price_unit}, adding: {partner_data['total_0']}, new total: {new_total}")
                    
                    existing_line_0.with_context(skip_date_check=True).write({
                        'price_unit': new_total,
                        'x_zlavnena_cena': new_total,
                        'date_start': contract.recurring_next_date,
                        'recurring_next_date': contract.recurring_next_date
                    })
                else:
                    _logger.info(f"Creating new 0% VAT line with amount: {partner_data['total_0']}")
                    self.env['contract.line'].with_context(skip_date_check=True).create({
                        'contract_id': contract.id,
                        'product_id': product_0.id,
                        'name': 'Vyúčtovanie paušálnych služieb a spotreby HLAS 0%',
                        'quantity': 1,
                        'price_unit': partner_data['total_0'],
                        'recurring_rule_type': 'monthly',
                        'recurring_interval': 1,
                        "uom_id": 1,
                        "x_zlavnena_cena": partner_data['total_0'],
                        'date_start': contract.recurring_next_date,
                        'recurring_next_date': contract.recurring_next_date,
                        'is_auto_renew': False,
                    })
                
        return True

    def action_reset_to_draft(self):
        """Reset invoice to draft"""
        self.write({'state': 'draft'})
        # Delete invoice lines
        self.invoice_line_ids.unlink()

    @api.model
    def _safe_convert_to_float(self, value):
        """Safely convert a string value to float, handling different formats."""
        if not value or pd.isna(value):
            return 0.0
            
        if isinstance(value, (int, float)):
            return float(value)
            
        # Clean the string
        value_str = str(value).strip()
        
        # Try direct conversion first
        try:
            return float(value_str)
        except ValueError:
            # Try replacing comma with period
            try:
                return float(value_str.replace(',', '.'))
            except ValueError:
                # Try replacing spaces and other characters
                try:
                    # Remove any currency symbols and other non-numeric characters except decimal separators
                    cleaned_value = re.sub(r'[^\d.,]', '', value_str)
                    # Replace comma with period for decimal point
                    cleaned_value = cleaned_value.replace(',', '.')
                    # If there are multiple decimal points, keep only the last one
                    if cleaned_value.count('.') > 1:
                        parts = cleaned_value.split('.')
                        cleaned_value = ''.join(parts[:-1]) + '.' + parts[-1]
                    return float(cleaned_value)
                except ValueError as e:
                    _logger.error(f"Failed to convert value '{value_str}' to float: {str(e)}")
                    return 0.0

    @api.model
    def _clean_phone_number(self, phone_number):
        """Remove all non-numeric characters from a phone number and strip leading '00'"""
        if not phone_number:
            return ''
        # Remove all non-numeric characters
        cleaned_number = re.sub(r'\D', '', str(phone_number))
        # Strip leading '00'
        return cleaned_number.lstrip('00')
    
    @api.model
    def reset_excess_usage_lines(self):
        _logger.info("Starting monthly reset of excess usage contract lines")
        
        # Find the product for excess usage
        product_0 = self.env['product.product'].search([
            ('display_name', '=', 'Vyúčtovanie paušálnych služieb a spotreby HLAS 0%')
        ], limit=1)
        product_23 = self.env['product.product'].search([
            ('display_name', '=', 'Vyúčtovanie paušálnych služieb a spotreby HLAS 23%')
        ], limit=1)

        if not product_0:
            _logger.error("Could not find product 'Vyúčtovanie paušálnych služieb a spotreby HLAS 0%'")
            return
        if not product_23:
            _logger.error("Could not find product 'Vyúčtovanie paušálnych služieb a spotreby HLAS 23%'")
            return            
        # Find all contract lines with this product
        contract_lines_0 = self.env['contract.line'].search([
            ('product_id', '=', product_0.id)
        ])
        contract_lines_23 = self.env['contract.line'].search([
            ('product_id', '=', product_23.id)
        ])


        reset_count = 0
        for line in contract_lines_0 + contract_lines_23:
            try:
                # Reset the line amount to 0
                line.write({
                    'price_unit': 0.0,
                    'x_zlavnena_cena': 0.0
                })
                reset_count += 1
            except Exception as e:
                _logger.error(f"Error resetting contract line {line.id}: {str(e)}")
                
        _logger.info(f"Successfully reset {reset_count} excess usage contract lines")

class ContractMobileInvoiceLine(models.Model):
    _name = "contract.mobile.invoice.line"
    _description = "Mobile Service Invoice Line"

    invoice_id = fields.Many2one(
        comodel_name="contract.mobile.invoice",
        string="Invoice",
        required=True,
        ondelete='cascade',
    )
    phone_number = fields.Char(string="Phone Number", required=True)
    service_name = fields.Char(string="Service Name")
    service_type = fields.Selection(
        selection=[
            ('basic', 'Basic Plan'),
            ('recurring', 'Recurring Service'),
            ('data', 'Data Usage'),
            ('voice', 'Voice Usage'),
            ('sms', 'SMS Usage'),
            ('mms', 'MMS Usage'),
            ('roaming', 'Roaming'),
            ('other', 'Other'),
        ],
        string="Service Type",
        default='other',
    )
    amount = fields.Float(string="Amount", digits=(16, 6))
    quantity = fields.Float(string="Quantity", digits=(16, 6))
    unit = fields.Char(string="Unit")
    total = fields.Float(string="Total", digits=(16, 2))
    is_excess_usage = fields.Boolean(string="Is Excess Usage", default=False)
    vat = fields.Char(string="DPH") # TODO: Add VAT to the invoice line
    mobile_service_id = fields.Many2one(
        comodel_name="contract.mobile.service",
        string="Mobile Service",
        compute="_compute_mobile_service",
        store=True,
    )
    
    partner_id = fields.Many2one(
        related="mobile_service_id.partner_id",
        string="Partner",
        store=True,
    )
    
    @api.depends('phone_number')
    def _compute_mobile_service(self):
        """Find the related mobile service based on phone number"""
        for record in self:
            # Clean phone number to ensure consistent format
            cleaned_number = self._clean_phone_number(record.phone_number)
            # Search for mobile service with matching number
            mobile_service = self.env['contract.mobile.service'].search([
                ('phone_number', 'like', cleaned_number),
                ('is_active', '=', True),
            ], limit=1)
            record.mobile_service_id = mobile_service.id if mobile_service else False
    
    @api.model
    def _clean_phone_number(self, phone_number):
        """Remove all non-numeric characters from a phone number and strip leading '00'"""
        if not phone_number:
            return ''
        # Remove all non-numeric characters
        cleaned_number = re.sub(r'\D', '', str(phone_number))
        # Strip leading '00'
        return cleaned_number.lstrip('00')


class ContractMobileUsageReport(models.Model):
    _name = "contract.mobile.usage.report"
    _description = "Mobile Usage Report"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "date desc"

    name = fields.Char(string="Name", required=True, tracking=True)
    date = fields.Date(string="Date", required=True, tracking=True)
    partner_id = fields.Many2one(
        comodel_name="res.partner",
        string="Partner",
        required=True,
        tracking=True,
    )
    invoice_ids = fields.Many2many(
        comodel_name="contract.mobile.invoice",
        string="Invoices",
        required=True,
        tracking=True,
    )
    # Removed operator_ids field and compute method as operator grouping is no longer needed
    report_line_ids = fields.One2many(
        comodel_name="contract.mobile.usage.report.line",
        inverse_name="report_id",
        string="Report Lines",
    )
    state = fields.Selection(
        selection=[
            ('draft', 'Draft'),
            ('done', 'Done'),
        ],
        string="State",
        default='draft',
        tracking=True,
    )
    company_id = fields.Many2one(
        comodel_name="res.company",
        string="Company",
        default=lambda self: self.env.company,
    )
    report_file = fields.Binary(string="Report File", attachment=True)
    report_filename = fields.Char(string="Report File Name")
    
    def action_generate_report(self):
        """Generate usage reports for all partners with mobile services across multiple invoices"""
        self.ensure_one()
        
        if not self.invoice_ids:
            raise UserError(_("No invoices selected for report generation"))
            
        try:
            # Get all partners with invoice lines across all invoices
            partner_ids = self.env['res.partner']
            for invoice in self.invoice_ids:
                partner_ids |= invoice.invoice_line_ids.mapped('partner_id')
            _logger.info(f"Found {len(partner_ids)} unique partners with mobile services")
            
            for partner in partner_ids:
                try:
                    # Find the mobilky contract for this partner
                    contract = self.env['contract.contract'].search([
                        ('partner_id', '=', partner.id),
                        ('x_contract_type', '=', 'Mobilky')
                    ], limit=1)
                    
                    if not contract:
                        _logger.warning(f"No Mobilky contract found for partner {partner.name}")
                        continue

                    _logger.info(f"Processing report for partner: {partner.name}")
                    
                    # Delete existing reports for this contract
                    existing_reports = self.env['ir.attachment'].search([
                        ('res_model', '=', 'contract.contract'),
                        ('res_id', '=', contract.id),
                        ('mimetype', '=', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    ])
                    if existing_reports:
                        _logger.info(f"Deleting {len(existing_reports)} existing reports for contract {contract.name}")
                        existing_reports.unlink()
                    
                    # Create report record for this partner
                    report = self.create({
                        'name': f"{self.name}_{partner.name}",
                        'date': self.date,
                        'partner_id': partner.id,
                        'invoice_ids': [(6, 0, self.invoice_ids.ids)],
                        'company_id': self.company_id.id,
                    })
                    
                    # Get invoice lines for this partner across all invoices
                    invoice_lines = self.env['contract.mobile.invoice.line']
                    for invoice in self.invoice_ids:
                        invoice_lines |= invoice.invoice_line_ids.filtered(
                            lambda l: l.partner_id == partner
                        )
                    
                    # Group lines by phone number
                    phone_groups = {}
                    for line in invoice_lines:
                        phone_key = line.phone_number
                        if phone_key not in phone_groups:
                            phone_groups[phone_key] = {
                                'mobile_service_id': line.mobile_service_id.id if line.mobile_service_id else False,
                                'partner_name': partner.name,
                                'phone_number': line.phone_number,
                                'basic_plan': '',
                                'basic_plan_cost': 0.0,
                                'excess_usage_cost': 0.0,
                                'total_cost': 0.0,
                                'total_data_usage': 0.0,  # Track total data usage
                                'total_sms_mms_usage': 0.0,  # Track total SMS/MMS count
                                'total_call_usage': 0.0,  # Track total call duration
                                'excess_data_usage': 0.0,
                                'excess_voice_usage': 0.0,
                                'excess_sms_usage': 0.0,
                                'is_company': partner.is_company,
                            }
                        
                        # Update group data based on line type
                        data = phone_groups[line.phone_number]
                        if line.service_type == 'basic':
                            data['basic_plan'] = line.service_name
                            data['basic_plan_cost'] = line.total
                        elif line.service_type == 'data':
                            data['total_data_usage'] += line.quantity  # Track ALL data usage
                            if line.is_excess_usage:
                                data['excess_data_usage'] += line.quantity
                        elif line.service_type == 'voice':
                            data['total_call_usage'] += line.quantity  # Track ALL call duration
                            if line.is_excess_usage:
                                data['excess_voice_usage'] += line.quantity
                        elif line.service_type in ['sms', 'mms']:
                            data['total_sms_mms_usage'] += line.quantity  # Track ALL SMS/MMS
                            if line.is_excess_usage:
                                data['excess_sms_usage'] += line.quantity
                        
                        if line.is_excess_usage:
                            data['excess_usage_cost'] += line.total
                        data['total_cost'] += line.total
                    
                    # Process each phone number
                    for phone_number, data in phone_groups.items():
                        # Create report line
                        self.env['contract.mobile.usage.report.line'].create({
                            'report_id': report.id,
                            'phone_number': phone_number,
                            'mobile_service_id': data['mobile_service_id'],
                            'partner_name': data['partner_name'],
                            'basic_plan': data['basic_plan'],
                            'basic_plan_cost': data['basic_plan_cost'],
                            'excess_usage_cost': data['excess_usage_cost'],
                            'total_cost': data['total_cost'],
                            'total_data_usage': data['total_data_usage'],  # Add total data usage
                            'total_sms_mms_usage': data['total_sms_mms_usage'],  # Add total SMS/MMS usage
                            'total_call_usage': data['total_call_usage'],  # Add total call usage
                            'excess_data_usage': data['excess_data_usage'],
                            'excess_voice_usage': data['excess_voice_usage'],
                            'excess_sms_usage': data['excess_sms_usage'],
                        })
                    
                    # Generate Excel report
                    report_content = report._generate_excel_report(phone_groups)
                    if report_content:
                        # Update report record with file
                        report.write({
                            'report_file': report_content,
                            'report_filename': f"{report.name}_{report.date}.xlsx",
                            'state': 'done'
                        })
                        
                        # Create new attachment for the contract
                        attachment = self.env['ir.attachment'].create({
                            'name': f"Výpis spotreby {report.report_filename.split('_')[1]}",
                            'type': 'binary',
                            'datas': report_content,
                            'res_model': 'contract.contract',
                            'res_id': contract.id,
                            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            'description': f"Report generovaný dňa {fields.Datetime.now() + timedelta(hours=2)}"
                        })
                        
                        _logger.info(f"Successfully generated and attached new report for partner {partner.name}")
                    
                except Exception as e:
                    _logger.error(f"Error processing partner {partner.name}: {str(e)}")
                    continue
            
            return True
            
        except Exception as e:
            _logger.error(f"Error generating reports: {str(e)}")
            raise UserError(_("Error generating reports: %s") % str(e))
            
        return False

    def _generate_excel_report(self, phone_groups):
        """Generate an Excel report file for each partner, containing all their phone numbers.
        Uses the same formatting and sections as nadspotreba.py."""
        # Get all invoice lines for the partner across all invoices
        invoice_lines = self.env['contract.mobile.invoice.line']
        for invoice in self.invoice_ids:
            invoice_lines |= invoice.invoice_line_ids.filtered(
                lambda l: l.partner_id == self.partner_id
            )
        try:
            base_path = os.path.dirname(os.path.abspath(__file__))
            novem_logo = os.path.join(base_path, '..', 'novem.png')
            
            # Group phone numbers by partner
            partner_groups = {}
            for phone_number, data in phone_groups.items():
                # Format phone number consistently
                formatted_phone = format_phone_number(phone_number)
                partner_name = data.get('partner_name', 'Unknown')
                partner_type = 'company' if data.get('is_company') else 'individual'
                
                key = (partner_type, partner_name)
                if key not in partner_groups:
                    partner_groups[key] = []
                # Store the formatted phone number
                data['formatted_phone'] = formatted_phone
                _logger.info(f"Processing phone number: {formatted_phone} for partner: {partner_name} ({partner_type})")
                # Format basic plan name
                data['formatted_plan'] = format_plan_name(data.get('basic_plan', ''))
                partner_groups[key].append((formatted_phone, data))

            # Continue with workbook creation...
            wb = Workbook()
            ws = wb.active
            
            # Add NOVEM logo if exists
            if os.path.exists(novem_logo):
                insert_image(ws, novem_logo, "A1")
            
            current_row = 5  # Start after logo
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Process each partner's data
            for partner_info, partner_data in partner_groups.items():
                for phone_number, data in sorted(partner_data):  # Sort by phone number
                    # Header section
                    ws.cell(row=current_row, column=1, value="Rozpis spotreby pre telefónne číslo:").font = Font(bold=True, color="438EAC")
                    ws.cell(row=current_row, column=2, value=data['formatted_phone'])  # Use formatted phone number
                    current_row += 1

                    # Basic plan section
                    _logger.info(f"Adding basic plan: {data['formatted_plan']}")
                    ws.cell(row=current_row, column=1, value=data['formatted_plan'])  # Use formatted plan name
                    current_row += 2
                    
                    # Rozpis účtovaných poplatkov
                    # Get all excessive usage lines for this phone number
                    excessive_lines = [line for line in invoice_lines if 
                                    line.phone_number == phone_number and 
                                    line.is_excess_usage and 
                                    line.total > 0]
                    
                    if excessive_lines:
                        ws.cell(row=current_row, column=1, value="Rozpis účtovaných poplatkov:").font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value="DPH").font = Font(bold=True)
                        ws.cell(row=current_row, column=3, value="Suma v EUR bez DPH").font = Font(bold=True)
                        for cell in [ws.cell(row=current_row, column=i) for i in range(1, 4)]:
                            cell.border = thin_border
                        current_row += 1

                        # List each excessive usage line separately
                        for line in excessive_lines:
                            ws.cell(row=current_row, column=1, value=line.service_name)
                            ws.cell(row=current_row, column=2, value=line.vat if line.vat else "0%")
                            x = f"{float(line.total):.4f}".replace('.', ',')
                            cell = ws.cell(row=current_row, column=3, value=x)
                            cell.number_format = '0.0000'
                            for cell in [ws.cell(row=current_row, column=i) for i in range(1, 4)]:
                                cell.border = thin_border
                            current_row += 1
                        current_row += 1

                    # Rozpis SMS / MMS
                    # Get SMS/MMS, voice, and data usage lines for this phone number
                    sms_lines = [line for line in invoice_lines if 
                              line.phone_number == phone_number and 
                              line.service_type in ['sms', 'mms']]
                    voice_lines = [line for line in invoice_lines if 
                                line.phone_number == phone_number and 
                                line.service_type == 'voice']
                    data_lines = [line for line in invoice_lines if 
                               line.phone_number == phone_number and 
                               line.service_type == 'data']

                    # Rozpis SMS / MMS
                    if sms_lines:
                        ws.cell(row=current_row, column=1, value="Rozpis SMS / MMS:").font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value="Počet kusov").font = Font(bold=True)
                        for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                            cell.border = thin_border
                        current_row += 1
                        
                        for line in sms_lines:
                            ws.cell(row=current_row, column=1, value=line.service_name)
                            ws.cell(row=current_row, column=2, value=f"{int(line.quantity)} ks")
                            for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                                cell.border = thin_border
                            current_row += 1
                        current_row += 1

                    # Rozpis volaní
                    if voice_lines:
                        ws.cell(row=current_row, column=1, value="Rozpis volaní:").font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value="Trvanie hovorov").font = Font(bold=True)
                        for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                            cell.border = thin_border
                        current_row += 1
                        
                        for line in voice_lines:
                            ws.cell(row=current_row, column=1, value=line.service_name)
                            duration = format_duration(line.quantity)
                            ws.cell(row=current_row, column=2, value=duration)
                            for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                                cell.border = thin_border
                            current_row += 1
                        current_row += 1

                    # Rozpis dát
                    if data_lines:
                        ws.cell(row=current_row, column=1, value="Rozpis dát:").font = Font(bold=True)
                        ws.cell(row=current_row, column=2, value="Spotreba dát").font = Font(bold=True)
                        for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                            cell.border = thin_border
                        current_row += 1
                        
                        for line in data_lines:
                            ws.cell(row=current_row, column=1, value=line.service_name)
                            formatted_usage = format_data_usage(line.quantity)
                            ws.cell(row=current_row, column=2, value=formatted_usage)
                            for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                                cell.border = thin_border
                            current_row += 1
                        current_row += 1

                    # Total section - only excess charges, not basic plan
                    excess_total = sum(line.total for line in invoice_lines if 
                                    line.phone_number == phone_number and 
                                    line.is_excess_usage and 
                                    line.service_type != 'basic')
                    
                    ws.cell(row=current_row, column=1, value=f"Faktúrovaná suma nad paušál bez DPH:").font = Font(bold=True)
                    x = f"{float(excess_total):.4f}".replace('.', ',')
                    cell = ws.cell(row=current_row, column=2, value=x)
                    cell.number_format = '0.0000'
                    for cell in [ws.cell(row=current_row, column=i) for i in range(1, 3)]:
                        cell.border = thin_border
                    current_row += 3  # Extra space between phone numbers

                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 100)
                    ws.column_dimensions[column].width = adjusted_width

                # Save workbook to memory
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                report_content = base64.b64encode(output.getvalue())
                return report_content # Return the first report since we're generating for a specific partner

        except Exception as e:
            _logger.error(f"Error generating Excel report: {str(e)}")
            return False
    def _normalize_plan_name(self, plan_name):
        """Normalize plan name from e-Net to NOVEM format"""
        if not plan_name:
            return ''
        
        name = plan_name.strip()
        # Convert e-Net to NOVEM
        if "e-Net" in name:
            name = name.replace("e-Net", "NOVEM")
        
        # Handle minutes in name
        if "minút" in name:
            name = name.replace("minút", "")
        elif "minut" in name:
            name = name.replace("minut", "")
        
        # Handle nekonečno
        if "nekonečno" in name:
            # Remove nekonečno and keep the data part
            name = name.replace("nekonečno", "").strip()
            name = f"NOVEM {name}"
        
        return name

    def _get_plan_data_size(self, plan_name):
        """Extract data size in GB from plan name"""
        if not plan_name or 'bez dát' in plan_name:
            return 0
        
        try:
            # Extract number followed by GB
            match = re.search(r'(\d+(?:,\d+)?)\s*GB', plan_name)
            if match:
                # Replace comma with dot for proper float conversion
                size = float(match.group(1).replace(',', '.'))
                return size
            # Handle specific case for 0.5GB
            if '0,5GB' in plan_name:
                return 0.5
            return 0
        except:
            return 0

    def _get_next_recommended_plan(self, current_plan, avg_monthly_usage_gb):
        """
        Get the next recommended plan based on usage.
        Returns tuple (plan_name, plan_data_gb)
        """
        # Define available data plans (GB) and their names
        plans = [
            (0, 'bez dát'),
            (0.5, '0,5GB'),
            (6, '6GB'),
            (10, '10GB'),
            (20, '20GB'),
            (30, '30GB'),
            (50, '50GB')
        ]
        
        # Normalize current plan name
        current_plan = self._normalize_plan_name(current_plan)
        
        # Determine plan type (Fér/250/regular)
        prefix = "NOVEM "
        if "Fér" in current_plan:
            prefix = "NOVEM Fér "
        elif "150" in current_plan:
            prefix = "NOVEM 150 "
        elif "250" in current_plan:
            prefix = "NOVEM 250 "

        # Get current plan size
        current_size = self._get_plan_data_size(current_plan)
        
        # Don't recommend upgrade if usage is less than 90% of current plan
        if current_size > 0 and avg_monthly_usage_gb < (current_size * 0.9):
            return None, None

        # For small plans (0.5GB), only upgrade if usage is significantly higher
        if current_size <= 0.5 and avg_monthly_usage_gb < 2:
            next_size = 0.5 if current_size == 0 else 6
            # simply build the suffix yourself; no need for index()
            suffix = '0,5GB' if next_size == 0.5 else f'{int(next_size)}GB'
            return f"{prefix}{suffix}", next_size

        # Find next suitable plan
        for size, suffix in plans:
            if size > current_size and size >= avg_monthly_usage_gb * 1.1:  # Give 10% buffer
                return f"{prefix}{suffix}", size
                
        # If we're here, user is already on a suitable plan
        return None, None

    
    @api.model
    def check_service_usage_patterns(self):
        """
        Check the last 3 months of usage reports to identify users with consistently high data usage
        that might benefit from a different service plan.
        """
        try:
            # Prevent duplicate runs by checking last execution
            last_run = self.env['ir.config_parameter'].sudo().get_param('last_usage_pattern_check')
            if last_run:
                last_run_dt = fields.Datetime.from_string(last_run)
                if last_run_dt + timedelta(minutes=1) > fields.Datetime.now():
                    _logger.info("Skipping usage pattern check - already run within 1 minute")
                    return

            # Get reports for last 3 full months (shifted back by one month since we analyze previous month's data)
            today = fields.Date.today()
            end_date = today.replace(day=1) - timedelta(days=1)  # Last day of previous month
            start_date = end_date.replace(day=1)  # First day of previous month
            start_date = start_date - timedelta(days=start_date.day)  # Last day of month before
            start_date = start_date.replace(day=1)  # First day
            start_date = start_date - timedelta(days=start_date.day)  # Last day of month before
            start_date = start_date.replace(day=1)  # First day of the third month back
            
            recent_reports = self.env['contract.mobile.usage.report'].search([
                ('date', '>=', start_date),
                ('date', '<=', end_date),
                ('state', '=', 'done')
            ])

            # Group reports by partner
            partner_reports = {}
            for report in recent_reports:
                if report.partner_id not in partner_reports:
                    partner_reports[report.partner_id] = []
                partner_reports[report.partner_id].append(report)

            # Analyze patterns for each partner
            issues_found = []
            for partner, reports in partner_reports.items():
                if len(reports) < 2:  # Need at least 2 months of data
                    continue

                phone_patterns = {}
                for report in reports:
                    for line in report.report_line_ids:
                        if line.phone_number not in phone_patterns:
                            phone_patterns[line.phone_number] = []
                        phone_patterns[line.phone_number].append({
                            'date': report.date,
                            'data_usage': line.total_data_usage,  # MB
                            'basic_plan': line.basic_plan,
                            'mobile_service': line.mobile_service_id,
                            'total_cost': line.total_cost,
                            'sms_mms_usage': line.total_sms_mms_usage,   # count
                            'voice_usage': line.total_call_usage,        # seconds
                        })

                # Check patterns for each phone number
                for phone, usages in phone_patterns.items():
                    if len(usages) < 2:  # Need at least 2 months of data for this number
                        continue
                        
                    # Check if the number is still active in contract.mobile.service
                    active_service = self.env['contract.mobile.service'].search([
                        ('phone_number', '=', phone),
                        ('partner_id', '=', partner.id),
                        ('is_active', '=', True),
                        ('ignore_alert', '=', False),
                    ], limit=1)
                    
                    if not active_service:
                        _logger.info(f"Skipping phone {phone} - no longer active for partner {partner.name}")
                        continue

                    # Sort by date
                    usages.sort(key=lambda x: x['date'])
                    
                    # Analyze usage
                    high_usage_months = 0
                    total_usage_gb = 0.0

                    high_sms_months = 0
                    total_sms_count = 0

                    high_voice_months = 0
                    total_voice_seconds = 0

                    current_plan = usages[-1]['basic_plan']
                    current_plan_size = self._get_plan_data_size(current_plan)
                    _logger.info(f"Phone {phone}: Current plan: {current_plan}, Plan size: {current_plan_size} GB")

                    # Thresholds based on plan name
                    plan_lower = current_plan.lower()

                    # SMS/MMS limits
                    if "nekonečno" in plan_lower:
                        sms_limit = None
                    elif "fér" in plan_lower:
                        sms_limit = 40
                    elif "250" in plan_lower or "150" in plan_lower:
                        sms_limit = 100
                    else:
                        sms_limit = None  # default finite-plan limit

                    # Voice limits (in minutes/month)
                    voice_limit = None
                    if "fér" in plan_lower:
                        voice_limit = 50
                    elif "250" in plan_lower:
                        voice_limit = 250
                    elif "150" in plan_lower:
                        voice_limit = 150
                    # else: None -> no voice check

                    for usage in usages:
                        # Data
                        # Convert bytes to GB (1 GB = 1024^3 bytes)
                        data_usage_bytes = usage['data_usage']
                        
                        # Unit detection and conversion
                        unit = "unknown"
                        if data_usage_bytes > 100000000:  # If it's in bytes (> ~100MB in bytes)
                            gb_usage = data_usage_bytes / (1024**3)
                            unit = "bytes"
                        else:  # If it's already in MB
                            gb_usage = data_usage_bytes / 1024
                            unit = "MB"
                            
                        _logger.info(f"Phone {phone}, date {usage['date']}: Data usage detected as {unit}: {data_usage_bytes} -> {gb_usage:.2f} GB")
                        total_usage_gb += gb_usage

                        # High data month?
                        if current_plan_size > 0.5 and gb_usage > (current_plan_size * 0.9):
                            high_usage_months += 1
                        elif current_plan_size <= 0.5 and gb_usage > 2:
                            high_usage_months += 1

                        # SMS/MMS
                        sms_count = usage.get('sms_mms_usage', 0)
                        total_sms_count += sms_count
                        if sms_limit and sms_count > sms_limit:
                            high_sms_months += 1

                        # Voice (seconds -> minutes)
                        voice_seconds = usage.get('voice_usage', 0) or 0
                        total_voice_seconds += voice_seconds
                        if voice_limit:
                            voice_mins_this_month = voice_seconds / 60.0
                            if voice_mins_this_month > voice_limit:
                                high_voice_months += 1

                    # Aggregates
                    months_n = len(usages)

                    avg_monthly_usage_gb = total_usage_gb / months_n
                    max_usage = max(usages, key=lambda x: x['data_usage'])
                    max_data_usage = max_usage['data_usage']
                    
                    # Apply the same logic for max usage
                    if max_data_usage > 100000000:  # If it's in bytes (> ~100MB in bytes)
                        max_usage_gb = max_data_usage / (1024**3)
                    else:  # If it's already in MB
                        max_usage_gb = max_data_usage / 1024
                        
                    max_usage_month = max_usage['date'].strftime('%m/%Y')
                    
                    _logger.info(f"Phone {phone}: Raw data usage values: {[u['data_usage'] for u in usages]}")
                    _logger.info(f"Phone {phone}: Avg usage: {avg_monthly_usage_gb:.2f} GB, Max usage: {max_usage_gb:.2f} GB")

                    avg_monthly_sms = total_sms_count / months_n
                    max_sms = max(usages, key=lambda x: x.get('sms_mms_usage', 0))
                    max_sms_count = max_sms.get('sms_mms_usage', 0)
                    max_sms_month = max_sms['date'].strftime('%m/%Y')

                    avg_monthly_voice_mins = (total_voice_seconds / 60.0) / months_n if months_n else 0.0
                    max_voice = max(usages, key=lambda x: x.get('voice_usage', 0) or 0)
                    max_voice_mins = (max_voice.get('voice_usage', 0) or 0) / 60.0
                    max_voice_month = max_voice['date'].strftime('%m/%Y')

                    # Decision: upgrades/downgrades
                    if (high_usage_months >= 2) or (sms_limit and high_sms_months >= 2) or (voice_limit and high_voice_months >= 2):
                        # Recommend a higher data plan (data-based)
                        recommended_plan, recommended_size = self._get_next_recommended_plan(
                            current_plan,
                            avg_monthly_usage_gb
                        )

                        # If SMS overuse, force into NOVEM 250 family (keep data size)
                        sms_needs_upgrade = sms_limit and high_sms_months >= 2
                        if sms_needs_upgrade and "250" not in current_plan:
                            size_for_suffix = recommended_size or current_plan_size
                            size_suffix = f"{size_for_suffix}GB" if size_for_suffix else "bez dát"
                            recommended_plan = f"NOVEM 250 {size_suffix}"
                            recommended_size = size_for_suffix

                        # If Voice overuse, also force into appropriate family; preference: keep NOVEM 250 if already chosen
                        voice_needs_upgrade = voice_limit and high_voice_months >= 2
                        if voice_needs_upgrade and "250" not in (recommended_plan or current_plan):
                            # Align to NOVEM 250 (voice limit 250) while keeping/bumping data size as already recommended
                            size_for_suffix = recommended_size or current_plan_size
                            size_suffix = f"{size_for_suffix}GB" if size_for_suffix else "bez dát"
                            recommended_plan = f"NOVEM 250 {size_suffix}"
                            recommended_size = size_for_suffix

                        if recommended_plan:
                            issues_found.append({
                                'partner_name': partner.name,
                                'phone_number': phone,
                                'current_plan': handle_o2_service_name(current_plan),
                                'current_plan_size': current_plan_size,
                                'avg_monthly_usage_gb': avg_monthly_usage_gb,
                                'max_usage_gb': max_usage_gb,
                                'max_usage_month': max_usage_month,
                                'recommended_plan': recommended_plan,
                                'recommended_plan_size': recommended_size,
                                'months_analyzed': months_n,
                                'high_usage_months': high_usage_months,
                                # SMS fields
                                'avg_monthly_sms': avg_monthly_sms,
                                'max_sms_count': max_sms_count,
                                'max_sms_month': max_sms_month,
                                'high_sms_months': high_sms_months,
                                'sms_limit': sms_limit,
                                # Voice fields
                                'avg_monthly_voice_mins': avg_monthly_voice_mins,
                                'max_voice_mins': max_voice_mins,
                                'max_voice_month': max_voice_month,
                                'high_voice_months': high_voice_months,
                                'voice_limit': voice_limit,
                                'type': 'upgrade',
                            })

                    elif current_plan_size > 0.5:
                        # Possible downgrade only if current plan has more than 0.5 GB (data-based only, unchanged)
                        usage_percentage = (avg_monthly_usage_gb / current_plan_size) * 100
                        max_usage_percentage = (max_usage_gb / current_plan_size) * 100
                        
                        if usage_percentage < 30 and max_usage_percentage < 50:
                            # Define available plans in descending order
                            plans = [
                                (50, '50GB'), (30, '30GB'), (20, '20GB'),
                                (10, '10GB'), (6, '6GB'), (0.5, '0,5GB')
                            ]
                            
                            # Find the next plan down that would still comfortably handle max usage
                            next_lower_plan = None
                            next_lower_size = 0
                            
                            for size, suffix in plans:
                                if size < current_plan_size and size >= max_usage_gb * 1.5:  # 50% buffer
                                    prefix = ""
                                    if "Fér" in current_plan:
                                        prefix = "NOVEM Fér "
                                    elif "250" in current_plan:
                                        prefix = "NOVEM 250 "
                                    elif "150" in current_plan:
                                        prefix = "NOVEM 150 "
                                    else:
                                        prefix = "NOVEM "
                                    next_lower_plan = f"{prefix}{suffix}"
                                    next_lower_size = size
                                    break
                            
                            if next_lower_plan:
                                issues_found.append({
                                    'partner_name': partner.name,
                                    'phone_number': phone,
                                    'current_plan': handle_o2_service_name(current_plan),
                                    'current_plan_size': current_plan_size,
                                    'avg_monthly_usage_gb': avg_monthly_usage_gb,
                                    'max_usage_gb': max_usage_gb,
                                    'max_usage_month': max_usage_month,
                                    'recommended_plan': next_lower_plan,
                                    'recommended_plan_size': next_lower_size,
                                    'months_analyzed': months_n,
                                    'high_usage_months': high_usage_months,
                                    # SMS fields
                                    'avg_monthly_sms': avg_monthly_sms,
                                    'max_sms_count': max_sms_count,
                                    'max_sms_month': max_sms_month,
                                    'high_sms_months': high_sms_months,
                                    'sms_limit': sms_limit,
                                    # Voice fields
                                    'avg_monthly_voice_mins': avg_monthly_voice_mins,
                                    'max_voice_mins': max_voice_mins,
                                    'max_voice_month': max_voice_month,
                                    'high_voice_months': high_voice_months,
                                    'voice_limit': voice_limit,
                                    'type': 'downgrade',
                                    'recommended_size': next_lower_size
                                })
            
            # If issues found, send email report
            if issues_found:
                # Split issues into upgrades and downgrades
                upgrades = [i for i in issues_found if i['type'] == 'upgrade']
                downgrades = [i for i in issues_found if i['type'] == 'downgrade']

                # Common inline styles (email-safe)
                container_style = (
                    "font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, "
                    "Oxygen, Ubuntu, Cantarell, 'Helvetica Neue', Arial, 'Noto Sans', sans-serif;"
                    "line-height:1.5;color:#111827;"
                )
                h2_style = "margin:0 0 8px 0;font-size:18px;font-weight:700;color:#111827;"
                p_style = "margin:0 0 16px 0;font-size:14px;color:#374151;"
                section_title = "margin:24px 0 8px 0;font-size:16px;font-weight:700;"
                table_style = (
                    "width:100%;border-collapse:collapse;margin:8px 0 16px 0;"
                    "font-size:13px;table-layout:auto;"
                )
                th_style = (
                    "text-align:left;padding:8px 10px;border:1px solid #E5E7EB;"
                    "background:#F9FAFB;font-weight:600;white-space:nowrap;"
                )
                td_style = "padding:8px 10px;border:1px solid #E5E7EB;vertical-align:top;"

                def render_section(title_text, title_color, rows):
                    if not rows:
                        return ""

                    # Detect if this section contains any SMS/Voice-related info
                    has_sms = any(
                        r.get('avg_monthly_sms') is not None or r.get('max_sms_count') is not None
                        for r in rows
                    )
                    has_voice = any(
                        r.get('avg_monthly_voice_mins') is not None or r.get('max_voice_mins') is not None
                        for r in rows
                    )

                    # Build the header dynamically
                    header_cells = [
                        "<th style='{th}'>Partner</th>",
                        "<th style='{th}'>Telefónne číslo</th>",
                        "<th style='{th}'>Aktuálny plán</th>",
                        "<th style='{th}'>Aktuálne dáta (GB)</th>",
                        "<th style='{th}'>Priemerná mesačná spotreba (GB)</th>",
                        "<th style='{th}'>Najvyššia spotreba (GB)</th>",
                        "<th style='{th}'>Mesiac najvyššej spotreby</th>",
                    ]
                    if has_sms:
                        header_cells += [
                            "<th style='{th}'>Priemerný počet SMS/MMS</th>",
                            "<th style='{th}'>Najvyšší počet SMS/MMS</th>",
                            "<th style='{th}'>Mesiac najvyššieho počtu SMS/MMS</th>",
                        ]
                    if has_voice:
                        header_cells += [
                            "<th style='{th}'>Priemerné minúty hovoru</th>",
                            "<th style='{th}'>Najviac minút hovoru</th>",
                            "<th style='{th}'>Mesiac najvyššieho počtu minút</th>",
                        ]
                    header_cells += [
                        "<th style='{th}'>Odporúčaný plán</th>",
                        "<th style='{th}'>Odporúčaný dátový objem (GB)</th>",
                        "<th style='{th}'>Analyzované mesiace</th>",
                    ]
                    header = "<tr>" + "".join(c.format(th=th_style) for c in header_cells) + "</tr>"

                    # Build body rows
                    body_rows = []
                    for r in rows:
                        row_cells = [
                            f"<td style='{td_style}'>{r['partner_name']}</td>",
                            f"<td style='{td_style}'>{r['phone_number']}</td>",
                            f"<td style='{td_style}'>{r['current_plan']}</td>",
                            f"<td style='{td_style}'>{r['current_plan_size']:.1f}</td>",
                            f"<td style='{td_style}'>{r['avg_monthly_usage_gb']:.1f}</td>",
                            f"<td style='{td_style}'>{r['max_usage_gb']:.1f}</td>",
                            f"<td style='{td_style}'>{r['max_usage_month']}</td>",
                        ]
                        if has_sms:
                            row_cells += [
                                f"<td style='{td_style}'>{int(r['avg_monthly_sms']) if r.get('avg_monthly_sms') is not None else '-'}</td>",
                                f"<td style='{td_style}'>{int(r['max_sms_count']) if r.get('max_sms_count') is not None else '-'}</td>",
                                f"<td style='{td_style}'>{r.get('max_sms_month') or '-'}</td>",
                            ]
                        if has_voice:
                            row_cells += [
                                f"<td style='{td_style}'>{int(round(r.get('avg_monthly_voice_mins', 0))) if r.get('avg_monthly_voice_mins') is not None else '-'}</td>",
                                f"<td style='{td_style}'>{int(round(r.get('max_voice_mins', 0))) if r.get('max_voice_mins') is not None else '-'}</td>",
                                f"<td style='{td_style}'>{r.get('max_voice_month') or '-'}</td>",
                            ]
                        row_cells += [
                            f"<td style='{td_style}'>{r['recommended_plan']}</td>",
                            f"<td style='{td_style}'>{r['recommended_plan_size']:.1f}</td>",
                            f"<td style='{td_style}'>{r['months_analyzed']}</td>",
                        ]
                        body_rows.append("<tr>" + "".join(row_cells) + "</tr>")

                    # Put together
                    return (
                        f"<h3 style='{section_title}color:{title_color};'>{title_text}</h3>"
                        f"<table role='table' style='{table_style}'>"
                        f"<thead>{header}</thead>"
                        f"<tbody>{''.join(body_rows)}</tbody>"
                        f"</table>"
                    )

                date_from = start_date.strftime('%d.%m.%Y')
                today = fields.Date.today()
                last_day_of_month = today.replace(day=1) - timedelta(days=1)
                date_to = last_day_of_month.strftime('%d.%m.%Y')

                email_body = (
                    f"<div style='{container_style}'>"
                    f"<h2 style='{h2_style}'>Sledovanie nadspotreby {date_from} – {date_to}</h2>"
                    f"<p style='{p_style}'>Automatický prehľad odporúčaných úprav balíkov podľa spotreby za sledované obdobie.</p>"
                    f"{render_section('Odporúčané zvýšenie balíka', '#d35400', upgrades)}"
                    f"{render_section('Odporúčané zníženie balíka', '#27ae60', downgrades)}"
                    f"</div>"
                )

                # Define CSV headers
                headers = [
                    'Partner', 'Telefónne číslo', 'Aktuálny plán', 'Aktuálne dáta (GB)',
                    'Priemerná mesačná spotreba (GB)', 'Najvyššia spotreba (GB)',
                    'Mesiac najvyššej spotreby', 'Priemerný počet SMS/MMS',
                    'Najvyšší počet SMS/MMS', 'Mesiac najvyššieho počtu SMS/MMS',
                    'Priemerné minúty hovoru', 'Najviac minút hovoru',
                    'Mesiac najvyššieho počtu minút', 'Odporúčaný plán',
                    'Odporúčaný dátový objem (GB)', 'Analyzované mesiace',
                    'Typ odporúčania'
                ]

                def create_csv_content(issues):
                    csv_data = io.StringIO()
                    csv_writer = csv.writer(csv_data, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                    csv_writer.writerow(headers)
                    
                    for issue in issues:
                        row = [
                            issue['partner_name'],
                            issue['phone_number'],
                            issue['current_plan'],
                            f"{issue['current_plan_size']:.1f}",
                            f"{issue['avg_monthly_usage_gb']:.1f}",
                            f"{issue['max_usage_gb']:.1f}",
                            issue['max_usage_month'],
                            str(int(issue['avg_monthly_sms'])) if issue.get('avg_monthly_sms') is not None else '-',
                            str(int(issue['max_sms_count'])) if issue.get('max_sms_count') is not None else '-',
                            issue.get('max_sms_month', '-'),
                            str(int(round(issue['avg_monthly_voice_mins']))) if issue.get('avg_monthly_voice_mins') is not None else '-',
                            str(int(round(issue['max_voice_mins']))) if issue.get('max_voice_mins') is not None else '-',
                            issue.get('max_voice_month', '-'),
                            issue['recommended_plan'],
                            f"{issue['recommended_plan_size']:.1f}",
                            str(issue['months_analyzed']),
                            'Zvýšenie' if issue['type'] == 'upgrade' else 'Zníženie'
                        ]
                        csv_writer.writerow(row)
                    
                    return base64.b64encode(csv_data.getvalue().encode('utf-8-sig'))

                # Group issues by partner
                all_issues = upgrades + downgrades
                partner_issues = {}
                regular_issues = []
                attachments = []

                for issue in all_issues:
                    partner_name = issue['partner_name']
                    if partner_name not in partner_issues:
                        partner_issues[partner_name] = []
                    partner_issues[partner_name].append(issue)

                # Process each partner's issues
                for partner_name, issues in partner_issues.items():
                    if len(issues) >= 3:
                        # Create separate CSV for this partner
                        partner_csv = create_csv_content(issues)
                        safe_name = ''.join(c for c in partner_name if c.isalnum() or c in (' ', '-', '_')).strip()
                        attachments.append({
                            'name': f'nadspotreba_{safe_name}_{date_from}_{date_to}.csv',
                            'type': 'binary',
                            'datas': partner_csv,
                            'mimetype': 'text/csv'
                        })
                    else:
                        # Add to regular issues for main CSV
                        regular_issues.extend(issues)

                # Create main CSV if there are any regular issues
                if regular_issues:
                    main_csv = create_csv_content(regular_issues)
                    attachments.append({
                        'name': f'nadspotreba_{date_from}_{date_to}.csv',
                        'type': 'binary',
                        'datas': main_csv,
                        'mimetype': 'text/csv'
                    })

                # Send email with all attachments
                mail_values = {
                    'email_from': self.env.company.email,
                    #'email_to': 'obrunovsky7@gmail.com,oliver.brunovsky@novem.sk',
                    'email_to': 'obrunovsky7@gmail.com,oliver.brunovsky@novem.sk,tomas.juricek@novem.sk',
                    'subject': 'Sledovanie nadspotreby mobilných služieb',
                    'body_html': email_body,
                    'attachment_ids': [(0, 0, att) for att in attachments]
                }
                self.env['mail.mail'].create(mail_values).send()

                # Update last run timestamp
                self.env['ir.config_parameter'].sudo().set_param(
                    'last_usage_pattern_check', 
                    fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                )

        except Exception as e:
            _logger.error(f"Error in service usage pattern analysis: {str(e)}")


class ContractMobileUsageReportLine(models.Model):
    _name = "contract.mobile.usage.report.line"
    _description = "Mobile Usage Report Line"

    report_id = fields.Many2one(
        comodel_name="contract.mobile.usage.report",
        string="Report",
        required=True,
        ondelete='cascade',
        index=True,
    )
    phone_number = fields.Char(string="Phone Number", required=True)
    mobile_service_id = fields.Many2one(
        comodel_name="contract.mobile.service",
        string="Mobile Service",
    )
    partner_name = fields.Char(string="Partner Name")
    basic_plan = fields.Char(string="Basic Plan")
    basic_plan_cost = fields.Float(string="Basic Plan Cost", digits=(16, 2))
    excess_usage_cost = fields.Float(string="Excess Usage Cost", digits=(16, 2))
    total_cost = fields.Float(string="Total Cost", digits=(16, 2))
    total_data_usage = fields.Float(string="Total Data Usage", digits=(16, 6))  # New field for total data usage
    total_sms_mms_usage = fields.Float(string="Total SMS/MMS Usage", digits=(16, 0))  # New field for total SMS/MMS counti
    total_call_usage = fields.Float(string="Total Call Usage", digits=(16, 6))  # New field for total call duration in seconds
    excess_data_usage = fields.Float(string="Excess Data Usage", digits=(16, 6))
    excess_voice_usage = fields.Float(string="Excess Voice Usage", digits=(16, 6))
    excess_sms_usage = fields.Float(string="Excess SMS Usage", digits=(16, 0))


def format_duration(seconds):
    """Format duration in seconds to HH:MM:SS"""
    try:
        seconds = float(seconds)
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02}:{minutes:02}:{secs:02}"
    except Exception as e:
        _logger.error(f"Error formatting duration: {str(e)}")
        return "00:00:00"

# Note: Using the format_data_usage function defined at the top of the file
def format_data_usage_redundant(volume_mb):
    """This function is deprecated. Use the format_data_usage function at the top of the file instead."""
    return format_data_usage(volume_mb)

def format_phone_number(number):
    """Format phone number to 421xxxxxxxxx format"""
    if not number:
        return ''
    # Remove all non-numeric characters
    cleaned = re.sub(r'\D', '', str(number))
    # Remove leading '00'
    cleaned = cleaned.lstrip('00')
    # If number starts with '0', replace it with '421'
    if cleaned.startswith('0'):
        cleaned = '421' + cleaned[1:]
    # If number doesn't start with '421', add it
    if not cleaned.startswith('421'):
        cleaned = '421' + cleaned
    return cleaned

def handle_o2_service_name(name):
    if not name:
        return ''
    name = name.strip()
    if "e-Net" in name:
        name = name.replace("e-Net", "NOVEM")
    
    if "fér" in name:
        name = name.replace("fér", "Fér")
    
    if "nekonečno" in name:
        name = name.replace("nekonečno ", "")
    if "minút" in name:
        name = name.replace("minút", "")
    if "minut" in name:
        name = name.replace("minut", "")

    return name

def format_plan_name(text):
    """Convert T-Biznis plan names to NOVEM equivalents"""
    if not text:
        return ''
        
    replacements = {
        "T-Biznis Flex - Variant 1": "NOVEM 6GB",
        "T-Biznis Flex - Variant 10": "NOVEM 10GB",
        "T-Biznis Flex - Variant 11": "NOVEM 30GB",
        "T-Biznis Flex - Variant 2": "NOVEM Fér bez dát",
        "T-Biznis Flex - Variant 3": "NOVEM 20GB",
        "T-Biznis Flex - Variant 4": "NOVEM 50GB",
        "T-Biznis Flex - Variant 5": "NOVEM 250 0,5GB",
        "T-Biznis Flex - Variant 6": "NOVEM 0,5GB",
        "T-Biznis Flex - Variant 7": "NOVEM 250 30 GB",
        "T-Biznis Flex - Variant 8": "NOVEM 250 10 GB",
        "T-Biznis Flex - Variant 9": "NOVEM bez dát"
    }
    
    # First, try exact match to avoid partial replacements
    result = text.strip()
    if result in replacements:
        return replacements[result]
    
    # If no exact match, try partial match but ensure full word matching
    for old, new in replacements.items():
        if old in result:
            result = result.replace(old, new)
            # Clean up any trailing digits that might have been left
            result = re.sub(r'(\d+GB)\d+', r'\1', result)
    
    if "e-Net" in result:
        result = result.replace("e-Net", "NOVEM")
        result = result.replace("minút", "")
        result = result.replace("minut", "")
    
    return result


